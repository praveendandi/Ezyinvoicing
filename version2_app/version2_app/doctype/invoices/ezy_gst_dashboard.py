from pydoc import doc
from re import T
from venv import create
from webbrowser import get
import datetime
# from attr import fields
# from cv2 import sort
import frappe
import json
import requests
import os
import pandas as pd
import os
import os.path
import sys
import numpy as np
from frappe.utils import data as date_util
from frappe.utils import cstr
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from frappe.utils.background_jobs import enqueue
from openpyxl.styles import Color, PatternFill, Font, Fill, colors, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell import Cell
# from UliPlot.XLSX import auto_adjust_xlsx_column_width
# from invoice_reconciliations import invoicereconciliationcount
# from xlsxwriter import add_worksheet
from version2_app.version2_app.report.outward_supplies.outward_supplies import execute

@frappe.whitelist(allow_guest=True)
def getGSTR1DashboardDetails(year=None, month=None):
    try:
        get_b2b_tax_invoice_summaries = frappe.db.sql(
            """SELECT count(name) as count, round(sum(total_gst_amount),2) as tax_amount, round(sum(pms_invoice_summary_without_gst),2) as taxable_value, round(sum(sales_amount_after_tax),2) as before_gst, round(sum(igst_amount),2) as igst_amount, round(sum(cgst_amount),2) as cgst_amount, round(sum(sgst_amount),2) as sgst_amount, round((sum(total_central_cess_amount)+sum(total_state_cess_amount)),2) as cess,invoice_category, '{}' as invoice_type, (sum(other_charges)-sum(non_revenue_amount)) as other_charges, sum(non_revenue_amount) as non_revenue_amount from `tabInvoices` where invoice_category='Tax Invoice' and invoice_type='B2B' and sez=0 and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format('B2B', year, month), as_dict=1)
        get_b2c_tax_invoice_summaries = frappe.db.sql(
            """SELECT count(name) as count, round(sum(total_gst_amount),2) as tax_amount, round(sum(pms_invoice_summary_without_gst),2) as taxable_value, round(sum(sales_amount_after_tax),2) as before_gst, round(sum(igst_amount),2) as igst_amount, round(sum(cgst_amount),2) as cgst_amount, round(sum(sgst_amount),2) as sgst_amount, round((sum(total_central_cess_amount)+sum(total_state_cess_amount)),2) as cess ,invoice_category, '{}' as invoice_type, (sum(other_charges)-sum(non_revenue_amount)) as other_charges, sum(non_revenue_amount) as non_revenue_amount from `tabInvoices` where invoice_category='Tax Invoice' and invoice_type='B2C' and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format('B2C', year, month), as_dict=1)
        get_b2b_credit_debit_invoice_summaries = frappe.db.sql(
            """SELECT count(name) as count, round(sum(total_gst_amount),2) as tax_amount, round(sum(pms_invoice_summary_without_gst),2) as taxable_value, round(sum(sales_amount_after_tax),2) as before_gst, round(sum(igst_amount),2) as igst_amount, round(sum(cgst_amount),2) as cgst_amount, round(sum(sgst_amount),2) as sgst_amount, round((sum(total_central_cess_amount)+sum(total_state_cess_amount)),2) as cess,'credit-debit' as invoice_category, '{}' as invoice_type, (sum(other_charges)-sum(non_revenue_amount)) as other_charges, sum(non_revenue_amount) as non_revenue_amount from `tabInvoices` where invoice_category in ('Credit Invoice','Debit Invoice') and `tabInvoices`.irn_generated = 'Success' and sez=0 and invoice_type='B2B' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format('B2B', year, month), as_dict=1)
        get_b2c_credit_debit_invoice_summaries = frappe.db.sql(
            """SELECT count(name) as count, round(sum(total_gst_amount),2) as tax_amount, round(sum(pms_invoice_summary_without_gst),2) as taxable_value, round(sum(sales_amount_after_tax),2) as before_gst, round(sum(igst_amount),2) as igst_amount, round(sum(cgst_amount),2) as cgst_amount, round(sum(sgst_amount),2) as sgst_amount, round((sum(total_central_cess_amount)+sum(total_state_cess_amount)),2) as cess,'credit-debit' as invoice_category, '{}' as invoice_type, (sum(other_charges)-sum(non_revenue_amount)) as other_charges, sum(non_revenue_amount) as non_revenue_amount from `tabInvoices` where invoice_category in ('Credit Invoice','Debit Invoice') and invoice_type='B2C' and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format('B2C', year, month), as_dict=1)
        get_hsn_summary = frappe.db.sql(
            """SELECT count(`tabItems`.parent) as count, (sum(`tabItems`.cgst_amount)+sum(`tabItems`.sgst_amount)+sum(`tabItems`.igst_amount)) as tax_amount, sum(`tabItems`.item_value_after_gst) as before_gst,sum(`tabItems`.item_taxable_value) as taxable_value, sum(`tabItems`.igst_amount) as igst_amount, sum(`tabItems`.cgst_amount) as cgst_amount, sum(`tabItems`.sgst_amount) as sgst_amount, 'hsn-summary' as invoice_category, 0 as other_charges,(sum(`tabInvoices`.total_state_cess_amount)+sum(`tabInvoices`.total_central_cess_amount)) as cess, 0 as non_revenue_amount from `tabItems` INNER JOIN `tabInvoices` ON `tabItems`.parent = `tabInvoices`.invoice_number where `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format(year, month), as_dict=1)
        nil_rated_supplies = frappe.db.sql("""SELECT count(`tabItems`.name) as count, sum(item_taxable_value) as taxable_value, sum(`tabItems`.cgst_amount)+sum(`tabItems`.sgst_amount)+sum(`tabItems`.igst_amount) as tax_amount, sum(item_value_after_gst) as before_gst, sum(`tabItems`.igst_amount) as igst_amount, sum(`tabItems`.cgst_amount) as cgst_amount, sum(`tabItems`.sgst_amount) as sgst_amount,'Nil Rated Supplies' as invoice_type, 0 as other_charges, 0 as cess, 0 as non_revenue_amount from `tabInvoices` INNER JOIN `tabItems` ON `tabItems`.parent = `tabInvoices`.name where ((`tabItems`.gst_rate = 0 and `tabItems`.taxable = "Yes") or (`tabItems`.taxable = "No") or (`tabInvoices`.sez = 1 and `tabItems`.type = "Excempted")) and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)='{}' and MONTH(invoice_date)='{}'""".format(year, month), as_dict=1)
        get_sez_SEZWP = frappe.db.sql(
            """SELECT count(name) as count, round(sum(total_gst_amount),2) as tax_amount, round(sum(pms_invoice_summary_without_gst),2) as taxable_value, round(sum(sales_amount_after_tax),2) as before_gst, round(sum(igst_amount),2) as igst_amount, round(sum(cgst_amount),2) as cgst_amount, round(sum(sgst_amount),2) as sgst_amount, round((sum(total_central_cess_amount)+sum(total_state_cess_amount)),2) as cess,invoice_category, '{}' as invoice_type, (sum(other_charges)-sum(non_revenue_amount)) as other_charges, sum(non_revenue_amount) as non_revenue_amount from `tabInvoices` where sez = 1 and invoice_type='B2B' and suptyp = 'SEZWP' and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format('B2B', year, month), as_dict=1)
        get_sez_SEZWOP = frappe.db.sql(
            """SELECT count(name) as count, round(sum(total_gst_amount),2) as tax_amount, round(sum(pms_invoice_summary_without_gst),2) as taxable_value, round(sum(sales_amount_after_tax),2) as before_gst, round(sum(igst_amount),2) as igst_amount, round(sum(cgst_amount),2) as cgst_amount, round(sum(sgst_amount),2) as sgst_amount, invoice_category, round((sum(total_central_cess_amount)+sum(total_state_cess_amount)),2) as cess,'{}' as invoice_type, (sum(other_charges)-sum(non_revenue_amount)) as other_charges, sum(non_revenue_amount) as non_revenue_amount from `tabInvoices` where sez = 1 and invoice_type='B2B' and suptyp = 'SEZWOP' and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}""".format('B2B', year, month), as_dict=1)

        advance_received = {"count": 0, "tax_amount": 0, "before_gst": 0,
                            "taxable_value": 0, "igst_amount": 0, "cgst_amount": 0, "sgst_amount": 0, "invoice_category": "advance-received","other_charges":0,"cess":0, "non_revenue_amount":0}
        adjustment_of_advances = {"count": 0, "tax_amount": 0, "before_gst": 0,
                                  "taxable_value": 0, "igst_amount": 0, "cgst_amount": 0, "sgst_amount": 0, "invoice_category": "adjustment-of-advances","other_charges":0,"cess":0, "non_revenue_amount":0}
        total_data = {"tax_b2b": {k: (0 if v is None else v) for k, v in get_b2b_tax_invoice_summaries[0].items()},
                      "sez_with_payment": {k: (0 if v is None else v) for k, v in get_sez_SEZWP[0].items()},
                      "sez_without_payment": {k: (0 if v is None else v) for k, v in get_sez_SEZWOP[0].items()},
                      "tax_b2c": {k: (0 if v is None else v) for k, v in get_b2c_tax_invoice_summaries[0].items()},
                      "credit_b2b": {k: (0 if v is None else v) for k, v in get_b2b_credit_debit_invoice_summaries[0].items()},
                      "credit_b2c": {k: (0 if v is None else v) for k, v in get_b2c_credit_debit_invoice_summaries[0].items()},
                      "nil_rated_supplies": {k: (0 if v is None else v) for k, v in nil_rated_supplies[0].items()},
                      "advance_received": advance_received,
                      "adjustment_of_advances": adjustment_of_advances,
                      "get_hsn_summary": {k: (0 if v is None else v) for k, v in get_hsn_summary[0].items()}}
        return {"success": True, "data": total_data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("getGSTR1DashboardDetails",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def getInvoices(filters=[], limit_page_length=20, limit_start=0, month=None, year=None, export=False):
    try:
        if isinstance(filters, str):
            filters = json.loads(filters)
        start_date = year+'-'+month+"-01"
        end_date = str(date_util.get_last_day(start_date))
        filters = filters + \
            [['invoice_date', 'between', [start_date, end_date]],
                ["irn_generated", "=", "Success"]]

        invoice_summary = frappe.db.get_list("Invoices", filters=filters, fields=['COUNT(DISTINCT(gst_number)) as no_of_suppliers', 'COUNT(DISTINCT(name)) as no_of_invoices', 'SUM(pms_invoice_summary_without_gst) as total_taxable_value',
                                                                                  'SUM(total_gst_amount) as total_gst_amount', 'SUM(pms_invoice_summary) as total_invoices_amount', 'SUM(other_charges) as other_charges', 'sum(non_revenue_amount) as non_revenue_amount', 'SUM(igst_amount) as total_igst', 'SUM(sgst_amount) as total_sgst',
                                                                                  'SUM(cgst_amount) as total_cgst', '(sum(total_central_cess_amount+total_state_cess_amount)) as cess'])
        if export == False:
            invoice_data = frappe.db.get_list("Invoices", filters=filters, fields=[
                '*'], start=int(limit_start), page_length=int(limit_page_length))
        else:
            invoice_data = frappe.db.get_list(
                "Invoices", filters=filters, fields=['invoice_number as InvoiceNo', 'DATE_FORMAT(invoice_date, "%d-%m-%Y") as InvoiceDate', 'gst_number as GSTINofSupplier', 'legal_name as LegalName', 'invoice_type as InvoiceType', 'sales_amount_after_tax as InvoiceAmt', 'pms_invoice_summary_without_gst as BaseAmt','(other_charges-non_revenue_amount) as OtherCharges', 'non_revenue_amount as NonRevenueCharges','sgst_amount as SGST', 'cgst_amount as CGST', 'igst_amount as IGST', 'total_gst_amount as TotalGST', '(total_central_cess_amount+total_state_cess_amount) as CESS'], order_by='name')
        return {"success": True, "data": invoice_data, "summary": invoice_summary[0]}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("getInvoices", "line No:{}\n{}".format(
            exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def getGSTR1ReconciliationSummaryCount(filters=[], month=None, year=None, company=None):
    try:
        if isinstance(filters, str):
            filters = json.loads(filters)
        start_date = year+'-'+month+"-01"
        end_date = date_util.get_last_day(start_date)
        filters = filters + \
            [['invoice_date', 'Between', [start_date, end_date]],
                ["irn_generated", "=", "Success"]]
        invoice_list = frappe.db.get_list(
            "Invoices", filters=filters, as_list=1)
        invoice_list = list(sum(invoice_list, ()))
        gst_invoice_list = frappe.db.get_list(
            "GSTR One Saved Invoices", filters=[["invoice_number", "in", invoice_list]], as_list=1)
        gst_invoice_list = list(sum(gst_invoice_list, ()))
        if any("property" in sublist for sublist in filters):
            miscellaneous_gst_invoice_list = frappe.db.get_list(
                "GSTR One Saved Invoices", filters=[["invoice_number", "in", invoice_list]], as_list=1)
        matching = len(list(set(invoice_list).intersection(gst_invoice_list)))
        missing_in_einvoice = len(
            list(set(gst_invoice_list)-set(invoice_list)))
        missing_in_gst = len(list(set(invoice_list)-set(gst_invoice_list)))
        data = {"total_invoices": matching+missing_in_einvoice+missing_in_gst,
                "matching": matching,
                "missing_in_einvoice": missing_in_einvoice,
                "missing_in_gst": missing_in_gst}
        month_year = (month+year).lstrip('0')
        gstr_one_saved_filter = [["period", "=", month_year]]
        if any("company" in sublist for sublist in filters):
            company = [each[2] for each in filters if "company" in each]
            gstr_one_saved_filter += [["company", "=", company[0]]]
        last_reconciled_on = frappe.db.get_list("Gstr One Saved Details", filters=gstr_one_saved_filter, fields=[
                                                "creation"], order_by='creation desc')
        if len(last_reconciled_on) > 0:
            last_reconciled_on = last_reconciled_on[0]["creation"]
        else:
            last_reconciled_on = ""
        return {"success": True, "data": data, "last_reconciled_on": last_reconciled_on}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("getGSTR1ReconciliationSummaryCount",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def getHsnSummary(filters=[], limit_page_length=20, limit_start=0, month=None, year=None, export=False, start_date=None, end_date=None):
    try:
        if isinstance(filters, str):
            filters = json.loads(filters)
        sql_filters = ""
        if len(filters) > 0:
            sql_filters = " and " + \
                (' and '.join("{} {} '{}'".format(
                    value[0], value[1], value[2]) for value in filters))
        if not export:
            if start_date and end_date:
                get_hsn_summary = frappe.db.sql(
                    """SELECT `tabItems`.sac_code as 'sac_code', `tabItems`.gst_rate, `tabItems`.unit_of_measurement_description as uqc, `tabItems`.quantity as total_quantity, sum(`tabItems`.cgst_amount) as 'cgst_amount', sum(`tabItems`.sgst_amount) as 'sgst_amount', sum(`tabItems`.igst_amount) as 'igst_amount', sum(`tabItems`.state_cess_amount) as 'state_cess_amount',sum(`tabItems`.cess_amount) as 'central_cess_amount', vat as vat, (sum(`tabItems`.cgst_amount)+sum(`tabItems`.sgst_amount)+sum(`tabItems`.igst_amount)) as 'total_gst', sum(`tabItems`.item_value) as 'total_tax_amount', sum(`tabItems`.item_value_after_gst) as 'total_amount' from `tabItems` INNER JOIN `tabInvoices` ON `tabItems`.parent = `tabInvoices`.invoice_number where `tabInvoices`.irn_generated = 'Success' and invoice_date between '{}' and '{}'{} GROUP BY `tabItems`.sac_code, `tabItems`.gst_rate""".format(start_date, end_date, sql_filters), as_dict=1)
            else:
                get_hsn_summary = frappe.db.sql(
                    """SELECT `tabItems`.sac_code as sac_code, `tabItems`.gst_rate, `tabItems`.unit_of_measurement_description as uqc, `tabItems`.quantity as total_quantity, sum(`tabItems`.cgst_amount) as cgst_amount, sum(`tabItems`.sgst_amount) as sgst_amount, sum(`tabItems`.igst_amount) as igst_amount, sum(`tabItems`.state_cess_amount) as state_cess_amount,sum(`tabItems`.cess_amount) as central_cess_amount, vat as vat, (sum(`tabItems`.cgst_amount)+sum(`tabItems`.sgst_amount)+sum(`tabItems`.igst_amount)) as total_gst, sum(`tabItems`.item_value) as total_tax_amount, sum(`tabItems`.item_value_after_gst) as total_amount from `tabItems` INNER JOIN `tabInvoices` ON `tabItems`.parent = `tabInvoices`.invoice_number where `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={}{} GROUP BY `tabItems`.sac_code, `tabItems`.gst_rate""".format(year, month, sql_filters), as_dict=1)
        else:
            get_hsn_summary = frappe.db.sql(
                """SELECT `tabItems`.sac_code as sac_code, `tabItems`.gst_rate, `tabItems`.unit_of_measurement_description as uqc, `tabItems`.quantity as total_quantity, sum(`tabItems`.cgst_amount) as cgst_amount, sum(`tabItems`.sgst_amount) as sgst_amount, sum(`tabItems`.igst_amount) as igst_amount, sum(`tabItems`.state_cess_amount) as state_cess_amount,sum(`tabItems`.cess_amount) as central_cess_amount, vat as vat, (sum(`tabItems`.cgst_amount)+sum(`tabItems`.sgst_amount)+sum(`tabItems`.igst_amount)) as total_gst, sum(`tabItems`.item_value) as total_tax_amount, sum(`tabItems`.item_value_after_gst) as total_amount from `tabItems` INNER JOIN `tabInvoices` ON `tabItems`.parent = `tabInvoices`.invoice_number where `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)={} and MONTH(invoice_date)={} GROUP BY `tabItems`.sac_code, `tabItems`.gst_rate""".format(year, month), as_dict=1)
        return {"success": True, "data": get_hsn_summary}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("getHsnSummary", "line No:{}\n{}".format(
            exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def export_invoices(filters=[], month=None, year=None):
    try:
        if isinstance(filters, str):
            filters = json.loads(filters)
        start_date = year+'-'+month+"-01"
        end_date = date_util.get_last_day(start_date)
        filters = filters + \
            [['invoice_date', 'Between', [start_date, end_date]],
                ["irn_generated", "=", "Success"]]
        invoice_data = frappe.db.get_list("Invoices", filters=filters, fields=['invoice_number as InvoiceNo', 'invoice_date as InvoiceDate', 'gst_number as GSTINofSupplier', 'legal_name as LegalName', 'invoice_type as InvoiceType', 'sales_amount_after_tax as InvoiceAmt',
                                          "sales_amount_before_tax as TatalTaxableAmount", "other_charges as OtherChanges", "cgst_amount as CGST", "sgst_amount as SGST", "igst_amount as IGST", "total_gst_amount as TotalGST", "(total_central_cess_amount+total_state_cess_amount) as CESS", "ack_date as EInvoiceGenerationDate"], order_by='invoice_number asc')
        if len(invoice_data) > 0:
            company = frappe.get_last_doc("company")
            cwd = os.getcwd()
            site_name = cstr(frappe.local.site)
            file_path = cwd + "/" + site_name + "/public/files/invoice_export.xlsx"
            df = pd.DataFrame.from_records(invoice_data)
            df.to_excel(file_path, index=False)
            files_new = {"file": open(file_path, 'rb')}
            payload_new = {'is_private': 1, 'folder': 'Home'}
            file_response = requests.post(company.host+"api/method/upload_file", files=files_new,
                                          data=payload_new, verify=False).json()
            if "file_url" in file_response["message"].keys():
                os.remove(file_path)
                return {"success": True, "file_url": file_response["message"]["file_url"]}
            return {"success": False, "message": "something went wrong"}
        else:
            return {"success": False, "message": "no data found"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("export_invoices", "line No:{}\n{}".format(
            exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def export_workbook(month=None, year=None):
    try:
        total_data = {}
        company = frappe.get_last_doc("company")
        cwd = os.getcwd()
        site_name = cstr(frappe.local.site)
        datetime_object = datetime.datetime.strptime(month, "%m")
        month_name = datetime_object.strftime("%b")
        file_path = cwd + "/" + site_name + \
            "/public/files/GSTR-"+month_name+"-"+year+".xlsx"
        get_summary = getGSTR1DashboardDetails(year, month)
        if not get_summary["success"]:
            return get_summary
        df = pd.DataFrame(get_summary["data"])
        df = df.T
        df1 = df.rename(index={'tax_b2b': 'B2B', 'sez_with_payment': 'B2B-SEZWP', 'sez_without_payment': 'B2B-SEZWOP', 'tax_b2c': 'B2C', 'credit_b2b': 'Credit/ Debit note (Registered)', 'credit_b2c': 'Credit/ Debit note (Unregistered)',
                               'nil_rated_supplies': 'Nil Rated Supplies', 'advance_received': 'Advance Received', 'adjustment_of_advances': 'Adjustment of Advances'})
        update_df = df1.drop('get_hsn_summary')
        total = update_df.sum()
        total.name = "Total"
        df1 = update_df.append(total.transpose())
        print(get_summary["data"]["get_hsn_summary"],";;;;;;;;;;;;;;;;;;;;;;;")
        df2 = df1.append(pd.Series(get_summary["data"]["get_hsn_summary"], index=df1.columns, name='HSN Summary of Outward supply'))
        df3 = df2.append(pd.Series(get_summary["data"]["nil_rated_supplies"], index=df2.columns, name='Nil-Rated-Supplies'))
        df4 = df3[10:12]
        total2 = df4.sum()
        total2.name = "Total2"
        df5 = df3.append(total2.transpose())
        # df4 = df3[17:][['before_gst', 'taxable_value','igst_amount','cgst_amount','sgst_amount','tax_amount','other_charges']]
        # total = df4.sum()
        # total.name = "Total"
        # df5 = df4.append(total.transpose())
        # total = df3.iloc['get_hsn_summary','nil_rated_supplies'].sum()
        # cols = ['before_gst','taxable_value','igst_amount','cgst_amount','sgst_amount','tax_amount','other_charges']
        # df['sum'] = df.loc[0:3, cols].sum(axis=1)
        # print(total,"//////")
        summary_data = df5.to_dict('index')
        get_hsn_summary = getHsnSummary(month=month, year=year, export=True)
        if not get_hsn_summary["success"]:
            return get_hsn_summary
        get_nil_rated = nil_rated_supplies(month=month, year=year)
        if not get_nil_rated["success"]:
            return get_nil_rated
        get_count_sequence = document_sequence(month=month, year=year)
        if not get_count_sequence["success"]:
            return get_count_sequence
        filter_list = {"B2B": [["invoice_type", "=", "B2B"], ["invoice_category", "=", "Tax Invoice"], ["sez", "=", 0]], "B2B-SEZWP": [["invoice_type", "=", "B2B"], ["invoice_category", "=", "Tax Invoice"], ["sez", "=", 1], ["suptyp", "=", "SEZWP"]], "B2B-SEZWOP": [["invoice_type", "=", "B2B"], ["invoice_category", "=", "Tax Invoice"], ["sez", "=", 1], ["suptyp", "=", "SEZWOP"]], "B2C": [["invoice_type", "=", "B2C"]], "Credit-Debit-B2B": [["invoice_type", "=", "B2B"], ["invoice_category", "in", ["Debit Invoice", "Credit Invoice"]], ["sez", "=", 0]], "Credit-Debit-B2C": [["invoice_type", "=", "B2C"], [
            "invoice_category", "in", ["Debit Invoice", "Credit Invoice"]]]}
        for key, value in filter_list.items():
            get_invoices_data = getInvoices(
                filters=value, month=month, year=year, export=True)
            if not get_invoices_data["success"]:
                return get_invoices_data
            total_data.update({key: get_invoices_data["data"]})
        # dataframe = pd.read_excel()
        wb = Workbook()
        ws = wb.active
        ws = wb.create_sheet("Summary")
        ws.title = "Summary"
        ws.move_range("A1:A5", rows=1, cols=0)
        fields = ['before_gst', 'taxable_value', 'igst_amount',
                  'cgst_amount', 'sgst_amount', 'tax_amount', 'other_charges', "non_revenue_amount", 'cess']
        ws.append(["Particulars", "Invoice value", "Taxable value",
                   "IGST", "CGST", "SGST", "TOTAL TAX LIABILITY", "Other Charges", "Non Revenue Amount", "CESS", ])
        for key, value in summary_data.items():
            values = (value[k] for k in fields)
            values = list(values)
            values.insert(0, key)
            ws.append(values)
        # ws.move_range("A6:A17", rows=0, cols=0)
        for i in range(ws.min_row, ws.max_row):
            ws.row_dimensions[i].height = 15
        for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J"]:
            ws.column_dimensions[i].width = 30
        font = Font(name='Cambria', size=12, bold=True, color='00FFFFFF')
        blueFill = PatternFill(start_color='0B0B45',
                               end_color='0B0B45', fill_type='solid')
        number_format = "#,##0"
        alignment = Alignment(horizontal='center', vertical='top')
        border = Border(left=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'), top=Side(border_style='thin', color='00000000'),
                        bottom=Side(border_style='thin', color='00000000'), diagonal=Side(border_style='thin', color='00000000'), diagonal_direction=0, outline=Side(border_style='thin', color='00000000'),
                        vertical=Side(border_style='thin', color='00000000'), horizontal=Side(border_style='thin', color='00000000'))
        cols = ['A', 'B', 'C']
        for col in cols:
            for i in range(2, 5):
                cols = ws[f'{col}{i}']
                cols.fill = blueFill
                cols.font = font
                cols.border = border
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I', 'J']
        for col in cols:
            cols = ws[f'{col}6']
            cols.fill = blueFill
            cols.font = font
            cols.alignment = alignment
            cols.border = border
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I', 'J']
        for col in cols:
            for i in range(7, 20):
                cols = ws[f'{col}{i}']
                cols.number_format = number_format
                cols.font = Font(name='Cambria', size=12)
                cols.border = border
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I', 'J']
        for col in cols:
            for i in [16, 19]:
                cols = ws[f'{col}{i}']
                cols.font = Font(bold=True)
        ws = wb.worksheets[1]
        ws['A2'] = 'GST Number : '+company.gst_number
        ws['A3'] = 'Name of the client : '+company.legal_name
        ws['A4'] = "GSTR-1 Summary for the period {} {}".format(
            month_name, year)
        invoice_fields = ['InvoiceNo', 'InvoiceDate', 'GSTINofSupplier', 'LegalName', 'InvoiceType',
                          'InvoiceAmt', 'BaseAmt', 'OtherCharges', 'NonRevenueCharges','SGST', 'CGST', 'IGST', 'TotalGST', 'CESS']
        for key, value in total_data.items():
            ws = wb.create_sheet(key)
            ws.title = key
            for i in ['A', 'B', 'C', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws.column_dimensions[i].width = 20
            ws.column_dimensions['D'].width = 60
            ws.append(['Invoice No', 'Invoice Date', 'GSTIN of Supplier', 'Legal Name', 'Invoice Type',
                      'Invoice Amt', 'Base Amt', 'OtherCharges', 'Non Revenue Charges','SGST', 'CGST', 'IGST', 'TotalGST', 'CESS'])
            if len(value) == 0:
                value = [{'InvoiceNo': None, 'InvoiceDate': None, 'GSTINofSupplier': None, 'LegalName': None, 'InvoiceType': None,
                          'InvoiceAmt': None, 'BaseAmt': None, 'OtherCharges': None, "NonRevenueCharges": None,'SGST': None, 'CGST': None, 'IGST': None, 'TotalGST': None, 'CESS': None}]
            else:
                invoices_df = pd.DataFrame.from_records(value)
                total = invoices_df.sum(numeric_only=True, axis=0)
                total.name = "Total"
                invoices_df = invoices_df.append(total.transpose())
                invoices_df['InvoiceNo'] = invoices_df['InvoiceNo'].replace(
                    np.nan, "Total")
                value = invoices_df.to_dict('records')
            for product in value:
                values = (product[k] for k in invoice_fields)
                ws.append(values)
            if ws.max_row > 2:
                for cell in ws[str(ws.max_row)+":"+str(ws.max_row)]:
                    cell.font = Font(bold=True)

            # df = pd.DataFrame.from_records(value)
        nil_rated = {"Inter state Supplies to Registered person": {"Nil Rated": get_nil_rated["data"]["inter_state_nill_rated_register_person"], "Exempted": get_nil_rated["data"]["inter_state_excempted_register_person"], "Non-GST": get_nil_rated["data"]
                                                                   ["inter_state_nonregister_register_person"]}, "Inter state Supplies to Unregistered person": {"Nil Rated": get_nil_rated["data"]["inter_state_nill_rated_unregister_person"], "Exempted": get_nil_rated["data"]["inter_state_excempted_unregister_person"], "Non-GST": get_nil_rated["data"]["inter_state_nonregister_unregister_person"]},
                     "Intra state Supplies to Registered person": {"Nil Rated": get_nil_rated["data"]["intra_state_nill_rated_register_person"], "Exempted": get_nil_rated["data"]["intra_state_excempted_register_person"], "Non-GST": get_nil_rated["data"]["intra_state_nonregister_register_person"]},
                     "Intra state Supplies to Unregistered person": {"Nil Rated": get_nil_rated["data"]["intra_state_nill_rated_unregister_person"], "Exempted": get_nil_rated["data"]["intra_state_excempted_unregister_person"], "Non-GST": get_nil_rated["data"]["intra_state_nonregister_unregister_person"]}}
        ws = wb.create_sheet("Nil Rated")
        ws.title = "Nil Rated"
        ws.column_dimensions["A"].width = 40
        for i in ['B', 'C', 'D']:
            ws.column_dimensions[i].width = 20
        nilrated_fields = ["Nil Rated", "Exempted", "Non-GST"]
        ws.append(["", "Nil Rated", "Exempted", "Non-GST"])
        nil_rated_df = pd.DataFrame(nil_rated)
        nil_rated_df = nil_rated_df.T
        total = nil_rated_df.sum(numeric_only=True, axis=0)
        total.name = "Total"
        nil_rated_df = nil_rated_df.append(total.transpose())
        nil_rated_list = nil_rated_df.to_dict('index')
        for key, value in nil_rated_list.items():
            values = (value[k] for k in nilrated_fields)
            values = list(values)
            values.insert(0, key)
            ws.append(values)
        for cell in ws[str(ws.max_row)+":"+str(ws.max_row)]:
            cell.font = Font(bold=True)
        hsn_fields = ['sac_code', "gst_rate", "uqc", "total_quantity", "cgst_amount", "sgst_amount", "igst_amount",
                      "state_cess_amount", "central_cess_amount", "total_gst", "total_tax_amount", "total_amount"]
        ws = wb.create_sheet("HSN_Summary")
        ws.title = "HSN_Summary"
        for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[i].width = 15
        ws.append(["SAC Code", "Gst Rate", "UQC", "Total Quantity", "Cgst Amount", "Sgst Amount", "Igst Amount",
                  "State Cess Amount", "Central Cess Amount", "Total Gst", "Total Tax Amount", "Total Amount"])
        hsn_summary_df = pd.DataFrame.from_records(get_hsn_summary["data"])
        total = hsn_summary_df.sum(numeric_only=True, axis=0)
        total.name = "Total"
        hsn_summary_df = hsn_summary_df.append(total.transpose())
        print(len(hsn_summary_df))
        
        # hsn_summary_df['Sac_Code'] = hsn_summary_df['Sac_Code'].replace(
        #     np.nan, "Total")
        # invoices_df['InvoiceNo'] = invoices_df['InvoiceNo'].replace(
        #             np.nan, "Total")
        hsn_summary = hsn_summary_df.to_dict('records')
        for product in hsn_summary:
            values = (product[k] for k in hsn_fields)
            ws.append(values)
        for cell in ws[str(ws.max_row)+":"+str(ws.max_row)]:
            cell.font = Font(bold=True)
        if len(hsn_summary_df) > 1:
            ws.cell(row=ws.max_row, column=2).value = ""
            ws.cell(row=ws.max_row, column=1).value = "Total"
        outward = outward_supply(month=month, year=year)
        if "success" in outward:
            return outward
        outward_fields = outward[0]
        outward_data = outward[1]
        ws = wb.create_sheet("OutWard Supply")
        ws.title = "OutWard_Supply"
        for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L','M','N','O','P','Q','R','S','T','U','V','W','X','Z']:
            ws.column_dimensions[i].width = 15
        ws.append(outward_fields)
        outward_df = pd.DataFrame.from_records(outward_data)
        total = outward_df.sum(numeric_only=True, axis=0)
        total.name = "Total"
        outward_supply_df = outward_df.append(total.transpose())
        outwards_supply = outward_supply_df.to_dict('records')
        for each in outwards_supply:
            values = (each[k] for k in outward_fields)
            ws.append(values)
        for cell in ws[str(ws.max_row)+":"+str(ws.max_row)]:
            cell.font = Font(bold=True)
        if len(outward_supply_df) > 1:
            ws.cell(row=ws.max_row, column=2).value = ""
            ws.cell(row=ws.max_row, column=1).value = "Total"
        # ws = wb.create_sheet("Sequence")
        # ws.title = "Sequence"
        # for i in ['A', 'B', 'C', 'D', 'E']:
        #     ws.column_dimensions[i].width = 15
        # # print(get_count_sequence["data"])
        # ws.append(["Document","From","To","Success","Cancelled", "Error"])
        # document_fields = ["Document","From","To","Success","Cancelled", "Error"]
        # sequence_count = [{"Document":"Tax Invoice", "From":get_count_sequence["data"]["tax_invoice_from"], "To":get_count_sequence["data"]["tax_invoice_to"], "Success":get_count_sequence["data"]["tax_invoice_success_count"], "Cancelled":get_count_sequence["data"]["tax_invoice_cancelled_count"], "Error":get_count_sequence["data"]["tax_invoice_error_count"]},{"Document":"Credit Invoice", "From":get_count_sequence["data"]["credit_invoice_from"], "To":get_count_sequence["data"]["credit_invoice_to"], "Success":get_count_sequence["data"]["credit_invoice_success_count"], "Cancelled":get_count_sequence["data"]["credit_invoice_cancelled_count"], "Error":get_count_sequence["data"]["credit_invoice_error_count"]}]
        # for product in sequence_count:
        #     values = (product[k] for k in document_fields)
        #     ws.append(values)
        Sheet = wb['Sheet']
        wb.remove(Sheet)
        wb.save(file_path)
        # return True
        files_new = {"file": open(file_path, 'rb')}
        payload_new = {'is_private': 1, 'folder': 'Home'}
        file_response = requests.post(company.host+"api/method/upload_file", files=files_new,
                                      data=payload_new, verify=False).json()
        if "file_url" in file_response["message"].keys():
            os.remove(file_path)
            return {"success": True, "file_url": file_response["message"]["file_url"]}
        return {"success": False, "message": "something went wrong"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("export_workbook", "line No:{}\n{}".format(
            exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def nil_rated_supplies(month=None, year=None):
    try:
        start_date = year+'-'+month+"-01"
        end_date = date_util.get_last_day(start_date)
        company = frappe.get_last_doc("company")
        inter_state_nill_rated_register_person = frappe.db.get_list("Invoices", filters=[["Items", "gst_rate", "=", 0], ["Items", "taxable", "=", "Yes"], ["Invoices", "place_of_supply", "=", company.state_code], [
                                                                    "Invoices", "invoice_type", "=", "B2B"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        inter_state_excempted_register_person = frappe.db.get_list("Invoices", filters=[["Items", "type", "=", "Excempted"], ["Invoices", "sez", "=", 1], ["Invoices", "place_of_supply", "=", company.state_code], [
                                                                   "Invoices", "invoice_type", "=", "B2B"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        inter_state_nonregister_register_person = frappe.db.get_list("Invoices", filters=[["Items", "taxable", "=", "No"], ["Invoices", "place_of_supply", "=", company.state_code], ["Invoices", "invoice_type", "=", "B2B"], [
                                                                     "Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        inter_state_nill_rated_unregister_person = frappe.db.get_list("Invoices", filters=[["Items", "gst_rate", "=", 0], ["Items", "taxable", "=", "Yes"], ["Invoices", "place_of_supply", "=", company.state_code], [
                                                                      "Invoices", "invoice_type", "=", "B2C"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        inter_state_excempted_unregister_person = frappe.db.get_list("Invoices", filters=[["Items", "type", "=", "Excempted"], ["Invoices", "sez", "=", 1], ["Invoices", "place_of_supply", "=", company.state_code], [
                                                                     "Invoices", "invoice_type", "=", "B2C"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        inter_state_nonregister_unregister_person = frappe.db.get_list("Invoices", filters=[["Items", "taxable", "=", "No"], ["Invoices", "place_of_supply", "=", company.state_code], ["Invoices", "invoice_type", "=", "B2C"], [
                                                                       "Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        intra_state_nill_rated_register_person = frappe.db.get_list("Invoices", filters=[["Items", "gst_rate", "=", 0], ["Items", "taxable", "=", "Yes"], ["Invoices", "place_of_supply", "!=", company.state_code], [
                                                                    "Invoices", "invoice_type", "=", "B2B"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        intra_state_excempted_register_person = frappe.db.get_list("Invoices", filters=[["Items", "type", "=", "Excempted"], ["Invoices", "sez", "=", 1], ["Invoices", "place_of_supply", "!=", company.state_code], [
                                                                   "Invoices", "invoice_type", "=", "B2B"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        intra_state_nonregister_register_person = frappe.db.get_list("Invoices", filters=[["Items", "taxable", "=", "No"], ["Invoices", "place_of_supply", "!=", company.state_code], ["Invoices", "invoice_type", "=", "B2B"], [
                                                                     "Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        intra_state_nill_rated_unregister_person = frappe.db.get_list("Invoices", filters=[["Items", "gst_rate", "=", 0], ["Items", "taxable", "=", "Yes"], ["Invoices", "place_of_supply", "!=", company.state_code], [
                                                                      "Invoices", "invoice_type", "=", "B2C"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        intra_state_excempted_unregister_person = frappe.db.get_list("Invoices", filters=[["Items", "type", "=", "Excempted"], ["Invoices", "sez", "=", 1], ["Invoices", "place_of_supply", "!=", company.state_code], [
                                                                     "Invoices", "invoice_type", "=", "B2C"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        intra_state_nonregister_unregister_person = frappe.db.get_list("Invoices", filters=[["Items", "taxable", "=", "No"], ["Invoices", "place_of_supply", "!=", company.state_code], [
                                                                       "Invoices", "invoice_type", "=", "B2C"], ["Invoices", "irn_generated", "=", "Success"], ['invoice_date', 'Between', [start_date, end_date]]], fields=["sum(`tabItems`.item_value_after_gst) as item"])
        data = {"inter_state_nill_rated_register_person": inter_state_nill_rated_register_person[0]["item"] if inter_state_nill_rated_register_person[0]["item"] else 0,
                "inter_state_excempted_register_person": inter_state_excempted_register_person[0]["item"] if inter_state_excempted_register_person[0]["item"] else 0,
                "inter_state_nonregister_register_person": inter_state_nonregister_register_person[0]["item"] if inter_state_nonregister_register_person[0]["item"] else 0,
                "inter_state_nill_rated_unregister_person": inter_state_nill_rated_unregister_person[0]["item"] if inter_state_nill_rated_unregister_person[0]["item"] else 0,
                "inter_state_excempted_unregister_person": inter_state_excempted_unregister_person[0]["item"] if inter_state_excempted_unregister_person[0]["item"] else 0,
                "inter_state_nonregister_unregister_person": inter_state_nonregister_unregister_person[0]["item"] if inter_state_nonregister_unregister_person[0]["item"] else 0,
                "intra_state_nill_rated_register_person": intra_state_nill_rated_register_person[0]["item"] if intra_state_nill_rated_register_person[0]["item"] else 0,
                "intra_state_excempted_register_person": intra_state_excempted_register_person[0]["item"] if intra_state_excempted_register_person[0]["item"] else 0,
                "intra_state_nonregister_register_person": intra_state_nonregister_register_person[0]["item"] if intra_state_nonregister_register_person[0]["item"] else 0,
                "intra_state_nill_rated_unregister_person": intra_state_nill_rated_unregister_person[0]["item"] if intra_state_nill_rated_unregister_person[0]["item"] else 0,
                "intra_state_excempted_unregister_person": intra_state_excempted_unregister_person[0]["item"] if intra_state_excempted_unregister_person[0]["item"] else 0,
                "intra_state_nonregister_unregister_person": intra_state_nonregister_unregister_person[0]["item"] if intra_state_nonregister_unregister_person[0]["item"] else 0
                }
        return {"success": True, "data": data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("nil_rated_supplies",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def nill_rated_items(month=None, year=None, limit_page_length=20, limit_start=0, Type="All"):
    try:
        count = 0
        if Type == "Non GST":
            filter = "`tabItems`.taxable = 'No'"
        elif Type == "Nil Rated":
            filter = "`tabItems`.gst_rate = 0 and `tabItems`.taxable = 'Yes'"
        elif Type == "Excempted":
            filter = "`tabInvoices`.sez = 1 and `tabItems`.type = 'Excempted'"
        else:
            filter = "((`tabItems`.gst_rate = 0 and `tabItems`.taxable = 'Yes') or (`tabItems`.taxable = 'No') or (`tabInvoices`.sez = 1 and `tabItems`.type = 'Excempted'))"
        nil_rated_supplies = frappe.db.sql("""SELECT invoice_number as 'Invoice_Number', item_name as 'Description', sac_code as 'SAC_Code',`tabItems`.item_value as 'Item_Value', `tabItems`.cgst_amount as CGST, `tabItems`.sgst_amount as SGST, `tabItems`.igst_amount as IGST, item_value_after_gst as 'Total_Value', if(`tabItems`.gst_rate = 0 and `tabItems`.taxable = "Yes", "Nil Rated", if(`tabItems`.taxable = 'No', 'Non GST', if(`tabInvoices`.sez = 1 and `tabItems`.type = "Excempted", "Excempted", ""))) as Type from `tabInvoices` INNER JOIN `tabItems` ON `tabItems`.parent = `tabInvoices`.name where {} and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)='{}' and MONTH(invoice_date)='{}' limit {}, {}""".format(filter, year, month, limit_start, limit_page_length), as_dict=1)
        supplies_count = frappe.db.sql(
            """SELECT count(`tabItems`.name) as count from `tabInvoices` INNER JOIN `tabItems` ON `tabItems`.parent = `tabInvoices`.name where {} and `tabInvoices`.irn_generated = 'Success' and YEAR(invoice_date)='{}' and MONTH(invoice_date)='{}'""".format(filter, year, month), as_dict=1)
        if len(supplies_count) > 0:
            count = supplies_count[0]["count"]
        return {"success": True, "data": nil_rated_supplies, "count": count}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("nill_rated_items",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def document_sequence(month=None, year=None):
    try:
        data = {}
        start_date = year+'-'+month+"-01"
        end_date = date_util.get_last_day(start_date)
        invoice_filters = {"tax_invoice_success_count": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "TAX INVOICE"], ["irn_generated", "=", "Success"]], "tax_invoice_cancelled_count": [["invoice_date", 'between', [start_date, end_date]], [
            "invoice_category", "=", "TAX INVOICE"], ["irn_generated", "=", "Cancelled"]], "tax_invoice_error_count": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "TAX INVOICE"], ["irn_generated", "=", "Error"]], "credit_invoice_success_count": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "CREDIT INVOICE"], ["irn_generated", "=", "Success"]], "credit_invoice_cancelled_count": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "CREDIT INVOICE"], ["irn_generated", "=", "Cancelled"]], "credit_invoice_error_count": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "CREDIT INVOICE"], ["irn_generated", "=", "Error"]]}
        total_invoice_list = []
        if bool(invoice_filters):
            for key, value in invoice_filters.items():
                invoice_numbers = frappe.db.get_list(
                    'Invoices', filters=value, pluck='name')
                data.update({key: len(invoice_numbers)})
                total_invoice_list.extend(invoice_numbers)
        tax_invoice_recon_data = frappe.db.get_list("Invoices", filters=[["invoice_date", 'between', [
                                                    start_date, end_date]], ["invoice_category", "=", "TAX INVOICE"]], pluck='name')
        credit_invoice_recon_data = frappe.db.get_list("Invoices", filters=[["invoice_date", 'between', [
                                                       start_date, end_date]], ["invoice_category", "=", "CREDIT INVOICE"]], pluck='name')
        if len(tax_invoice_recon_data) > 0:
            data["tax_invoice_from"] = min(tax_invoice_recon_data)
            data["tax_invoice_to"] = max(tax_invoice_recon_data)
        else:
            data["tax_invoice_from"] = ""
            data["tax_invoice_to"] = ""
        if len(credit_invoice_recon_data) > 0:
            data["credit_invoice_from"] = min(credit_invoice_recon_data)
            data["credit_invoice_to"] = max(credit_invoice_recon_data)
        else:
            data["credit_invoice_from"] = ""
            data["credit_invoice_to"] = ""
        invoices_list = tax_invoice_recon_data + credit_invoice_recon_data
        d110_total_invoices = frappe.db.get_list("Invoice Reconciliations", filters=[
            ["bill_generation_date", "between", [start_date, end_date]]], pluck="name")
        if len(d110_total_invoices) > 0 and len(invoices_list) > 0:
            data["missing_in_d110"] = list(
                set(invoices_list) - set(d110_total_invoices))
            data["missing_in_ezy_invoicing"] = list(
                set(d110_total_invoices) - set(invoices_list))
        else:
            data["tax_invoice_success_count"] = 0
            data["tax_invoice_cancelled_count"] = 0
            data["tax_invoice_error_count"] = 0
            data["credit_invoice_success_count"] = 0
            data["credit_invoice_cancelled_count"] = 0
            data["credit_invoice_error_count"] = 0
        return {"success": True, "data": data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("document_sequence",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def reconciliation(start_date, end_date, workbook="true"):
    try:
        # if isinstance(filters, str):
        #     filters = json.loads(filters)
        company = frappe.get_last_doc("company")
        cwd = os.getcwd()
        site_name = cstr(frappe.local.site)
        month_name = datetime.datetime.strptime(
            start_date, "%Y-%m-%d").strftime("%b")
        # month_name = datetime_object.strftime("%b")
        year_object = datetime.datetime.strptime(
            start_date, "%Y-%m-%d").strftime("%Y")
        file_path = cwd + "/" + site_name + \
            "/public/files/RECON-"+month_name+"-"+year_object+".xlsx"
        missing = get_missing_invoices(start_date, end_date)
        if not missing["success"]:
            return missing
        # return missing
        recon_data = {}
        sequence_data = [{"Ezyinvoicing Count": len(missing["data"]["ezy_invoicing_invoices"]), "Opera Folios Count": len(missing["data"]["opera_folios"]), "Missing In EzyInvoicing": len(missing["data"]["missing_in_ezyinvoicing"]), "Missing In Opera": len(
            missing["data"]["missing_in_opera"]), "Type Missmatch": len(missing["data"]["type_missmatch_for_tax"])+len(missing["data"]["type_missmatch_for_credit"]) if len(missing["data"]["type_missmatch_for_credit"]) > 1 else 0}]
        # Summary of Invoices
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        seq_data = [{k.replace(' ', '_'): v for k, v in d.items()}
                    for d in sequence_data]
        recon_data["count"] = seq_data
        df_summary = pd.DataFrame(sequence_data)
        df_summary.to_excel(writer, sheet_name="Count", index=False)
        for each in ["Ezyinvoicing Count", "Opera Folios Count", "Missing In EzyInvoicing", "Missing In Opera", "Type Missmatch"]:
            col_idx = df_summary.columns.get_loc(each)
            writer.sheets['Count'].set_column(col_idx, col_idx, 20)
        # Missing Invoices In Opera
        missing_in_opera = {
            "Invoices Numbers": missing["data"]["missing_in_opera"]}
        get_missing_in_opera_data = [{"Invoice Number": "", "Invoice Date": "", "Invoice Type": "", "Invoice Category": "", "IRN Status": "", "EzyInvoice Status": "", "Invoice Amount": "", "Taxable Value": "",
                                      "IGST Amount": "", "SGST Amount": "", "CGST Amount": "", "Total Gst Amount": "", "Other Charges": "", "CESS": "", "VAT Amount": "", "IRN Number": "", "Acknowledgement No": "", "Acknowledgement Date": ""}]
        if len(missing["data"]["missing_in_opera"]) > 0:
            # get_missing_in_opera_data = frappe.db.get_list("Invoices", filters=[["name","in",missing["data"]["missing_in_opera"]]],fields=['invoice_number as InvoiceNumber', 'invoice_date as InvoiceDate','invoice_type as InvoiceType', 'invoice_category as InvoiceCategory', 'irn_generated as IRNStatus', 'invoice_from as EzyInvoiceStatus', 'pms_invoice_summary_without_gst as TaxableValue', 'total_gst_amount as TotalGstAmount', 'pms_invoice_summary as InvoicesAmount', 'other_charges as OtherCharges', 'igst_amount as IGSTAmount', 'sgst_amount as SGSTAmount', 'cgst_amount as CGSTAmount', '(total_central_cess_amount+total_state_cess_amount) as CESS','total_vat_amount as VATAmount', 'If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as IRNNumber', 'ack_no as AcknowledgementNo', 'ack_date as AcknowledgementDate'], order_by='name')
            missing_numbers = tuple(missing["data"]["missing_in_opera"])
            if len(missing_numbers) == 1:
                missing_numbers = missing_numbers + ('0',)
            get_missing_in_opera = frappe.db.sql("""select invoice_number as 'Invoice Number', invoice_date as 'Invoice Date',invoice_type as 'Invoice Type', invoice_category as 'Invoice Category', irn_generated as 'IRN Status', invoice_from as 'EzyInvoice Status', pms_invoice_summary as 'Invoice Amount', pms_invoice_summary_without_gst as 'Taxable Value', igst_amount as 'IGST Amount', sgst_amount as 'SGST Amount', cgst_amount as 'CGST Amount', total_gst_amount as 'Total Gst Amount', other_charges as 'Other Charges', (total_central_cess_amount+total_state_cess_amount) as CESS,total_vat_amount as 'VAT Amount', If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as 'IRN Number',If(invoice_category!="Credit Invoice",ack_no,credit_ack_no) as 'Acknowledgement No',  If(invoice_category!="Credit Invoice",ack_date,credit_ack_date) as 'Acknowledgement Date' from `tabInvoices` where name in {} and irn_generated != 'Zero Invoice' ORDER BY name""".format(missing_numbers), as_dict=1)
            if len(get_missing_in_opera) > 0:
                get_missing_in_opera_data = get_missing_in_opera
        miss_in_opera = [{k.replace(' ', '_'): v for k, v in d.items()}
                         for d in get_missing_in_opera_data]
        recon_data["Missing_In_Opera"] = miss_in_opera
        df_missing_opera_folios_list = pd.DataFrame.from_records(
            get_missing_in_opera_data)
        df_missing_opera_folios_list.to_excel(
            writer, sheet_name="Missing In Opera", index=False)
        for each in ["Invoice Number", "Invoice Date", "Invoice Type", "Invoice Category", "IRN Status", "EzyInvoice Status", "Invoice Amount", "Taxable Value", "IGST Amount", "SGST Amount", "CGST Amount", "Total Gst Amount", "Other Charges", "CESS", "VAT Amount", "IRN Number", "Acknowledgement No", "Acknowledgement Date"]:
            col_idx = df_missing_opera_folios_list.columns.get_loc(each)
            writer.sheets['Missing In Opera'].set_column(col_idx, col_idx, 20)
        missing_in_opera = {
            "Invoices Numbers": missing["data"]["missing_in_opera"]}
        get_missing_in_opera_data = [{"Invoice Number": "", "Invoice Date": "", "Invoice Type": "", "Invoice Category": "", "IRN Status": "", "EzyInvoice Status": "", "Invoice Amount": "", "Taxable Value": "",
                                      "IGST Amount": "", "SGST Amount": "", "CGST Amount": "", "Total Gst Amount": "", "Other Charges": "", "CESS": "", "VAT Amount": "", "IRN Number": "", "Acknowledgement No": "", "Acknowledgement Date": ""}]
        if len(missing["data"]["missing_in_opera"]) > 0:
            # get_missing_in_opera_data = frappe.db.get_list("Invoices", filters=[["name","in",missing["data"]["missing_in_opera"]]],fields=['invoice_number as InvoiceNumber', 'invoice_date as InvoiceDate','invoice_type as InvoiceType', 'invoice_category as InvoiceCategory', 'irn_generated as IRNStatus', 'invoice_from as EzyInvoiceStatus', 'pms_invoice_summary_without_gst as TaxableValue', 'total_gst_amount as TotalGstAmount', 'pms_invoice_summary as InvoicesAmount', 'other_charges as OtherCharges', 'igst_amount as IGSTAmount', 'sgst_amount as SGSTAmount', 'cgst_amount as CGSTAmount', '(total_central_cess_amount+total_state_cess_amount) as CESS','total_vat_amount as VATAmount', 'If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as IRNNumber', 'ack_no as AcknowledgementNo', 'ack_date as AcknowledgementDate'], order_by='name')
            missing_numbers = tuple(missing["data"]["missing_in_opera"])
            if len(missing_numbers) == 1:
                missing_numbers = missing_numbers + ('0',)
            get_missing_in_opera = frappe.db.sql("""select invoice_number as 'Invoice Number', invoice_date as 'Invoice Date',invoice_type as 'Invoice Type', invoice_category as 'Invoice Category', irn_generated as 'IRN Status', invoice_from as 'EzyInvoice Status', pms_invoice_summary as 'Invoice Amount', pms_invoice_summary_without_gst as 'Taxable Value', igst_amount as 'IGST Amount', sgst_amount as 'SGST Amount', cgst_amount as 'CGST Amount', total_gst_amount as 'Total Gst Amount', other_charges as 'Other Charges', (total_central_cess_amount+total_state_cess_amount) as CESS,total_vat_amount as 'VAT Amount', If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as 'IRN Number',If(invoice_category!="Credit Invoice",ack_no,credit_ack_no) as 'Acknowledgement No',  If(invoice_category!="Credit Invoice",ack_date,credit_ack_date) as 'Acknowledgement Date' from `tabInvoices` where name in {} and irn_generated = 'Zero Invoice' ORDER BY name""".format(missing_numbers), as_dict=1)
            if len(get_missing_in_opera) > 0:
                get_missing_in_opera_data = get_missing_in_opera
        miss_in_opera = [{k.replace(' ', '_'): v for k, v in d.items()}
                         for d in get_missing_in_opera_data]
        recon_data["Zero_Invoices"] = miss_in_opera
        df_missing_opera_folios_list = pd.DataFrame.from_records(
            get_missing_in_opera_data)
        df_missing_opera_folios_list.to_excel(
            writer, sheet_name="Zero Invoices", index=False)
        for each in ["Invoice Number", "Invoice Date", "Invoice Type", "Invoice Category", "IRN Status", "EzyInvoice Status", "Invoice Amount", "Taxable Value", "IGST Amount", "SGST Amount", "CGST Amount", "Total Gst Amount", "Other Charges", "CESS", "VAT Amount", "IRN Number", "Acknowledgement No", "Acknowledgement Date"]:
            col_idx = df_missing_opera_folios_list.columns.get_loc(each)
            writer.sheets['Zero Invoices'].set_column(col_idx, col_idx, 20)
        # Missing Invoices In Ezyinvoicing
        missing_in_ezy = {
            "Invoices Numbers": missing["data"]["missing_in_ezyinvoicing"]}
        missing_in_ezyinvoicing = [
            {k.replace(' ', '_'): v for k, v in d.items()} for d in get_missing_in_opera_data]
        recon_data["Missing_In_EzyInvoicing"] = missing_in_ezyinvoicing
        df_missing_ezy_invoicing_list = pd.DataFrame.from_dict(missing_in_ezy)
        df_missing_ezy_invoicing_list.to_excel(
            writer, sheet_name="Missing In EzyInvoicing", index=False)
        for each in ["Invoices Numbers"]:
            col_idx = df_missing_ezy_invoicing_list.columns.get_loc(each)
            writer.sheets['Missing In EzyInvoicing'].set_column(
                col_idx, col_idx, 20)
        # Type Missmatch
        type_missmatch = missing["data"]["type_missmatch_for_tax"] + \
            missing["data"]["type_missmatch_for_credit"]
        type_mismatch = [
            {k.replace(' ', '_'): v for k, v in d.items()} for d in type_missmatch]
        recon_data["Invoice_Type_Missmatch"] = type_mismatch
        df_type_missmatch = pd.DataFrame.from_records(type_missmatch)
        print(df_type_missmatch, "....................")
        # frappe.db.get_list("Invoices", filters=filters, fields=['SUM(pms_invoice_summary_without_gst) as total_taxable_value', 'SUM(total_gst_amount) as total_gst_amount', 'SUM(pms_invoice_summary) as total_invoices_amount', 'SUM(other_charges) as other_charges', 'SUM(igst_amount) as total_igst', 'SUM(sgst_amount) as total_sgst', 'SUM(cgst_amount) as total_cgst', 'SUM(total_central_cess_amount+total_state_cess_amount) as cess'])
        # gst_invoice_list = frappe.db.get_list(
        #     "Missing In Opera", filters=[["invoice_number", "in", missing_in_opera]], as_list=1)
        df_type_missmatch.to_excel(
            writer, sheet_name="Invoice Type Missmatch", index=False)
        for each in ["InvoiceNumber", "Ezyinvoicing", "Opera"]:
            col_idx = df_type_missmatch.columns.get_loc(each)
            writer.sheets['Invoice Type Missmatch'].set_column(
                col_idx, col_idx, 20)
        # auto_adjust_xlsx_column_width(
        #     df_type_missmatch, writer, sheet_name="Invoice Type Missmatch", margin=0)
        comparing = invoices_summary(start_date, end_date)
        if not comparing["success"]:
            return comparing
        comparing1 = [{k.replace(' ', '_'): v for k, v in d.items()}
                      for d in comparing["data"]]
        recon_data["Ezy_Invoicing_Summary"] = comparing1
        df_comparing_totals = pd.DataFrame.from_records(comparing["data"])
        # df1 = df_comparing_totals
        # hsn = getHsnSummary(start_date=start_date, end_date=end_date)
        # if not hsn["success"]:
        #     return hsn
        # hsnb2b = [{k.replace(' ', '_') : v for k, v in d.items()} for d in hsn["data"]]
        # recon_data["HSN"] = hsnb2b
        # if len(hsn["data"]) == 0:
        #     hsn["data"] = [{"SAC Code": "", "Gst Rate": "", "UQC": "", "Total Quantity": "", "CGST Amount": "", "SGST Amount": "", "IGST Amount": "", "State CESS Amount": "", "Central CESS Amount": "", "VAT": "", "Total GST": "", "Total Tax Amount": "", "Total Amount": ""}]
        # df_b2b_hsn = pd.DataFrame.from_records(hsn["data"])
        # total = df_b2b_hsn.sum(numeric_only=True, axis=0)
        # total.name = "Total"
        # df_b2b = df_b2b_hsn.append(total.transpose())
        # if len(hsn["data"]) > 1:
        #     df_b2b['SAC Code'] = df_b2b['SAC Code'].replace(
        #                 np.nan, "Total")
        #     df_b2b['Gst Rate'] = np .where(df_b2b['Gst Rate'] > 25, "", df_b2b["Gst Rate"])
        # df2 = df_b2b
        # # df2 = pd.DataFrame({
        # # "A": ["A4", "A5", "A6", "A7","A8", "A9", "A10", "A11", "A12", "A13"],},
        # # index=[4,5,6,7,8,9,10,11,12,13],)
        # # result = pd.concat([df1, df2], ignore_index=True, sort=False)
        # pieces = {"x": df1, "y": df2}
        # result = pd.concat(pieces)
        # print(result, ";;;;;;;;;;;..............,,,,,,,,,,,,,,,,,,>>><<<<")
        # result.to_excel(
        #     writer, sheet_name="Ezy Invoicing Summary", index=False)
        # for each in ["SAC Code", "Gst Rate", "UQC", "Total Quantity", "CGST Amount", "SGST Amount", "IGST Amount", "State CESS Amount", "Central CESS Amount", "VAT", "Total GST", "Total Tax Amount", "Total Amount"]:
        #     col_idx = df_b2b.columns.get_loc(each)
        #     writer.sheets['Ezy Invoicing Summary'].set_column(col_idx, col_idx, 20)
        df_comparing_totals.to_excel(
            writer, sheet_name="Ezy Invoicing Summary", index=False)
        for each in ["BeforeTax", "CGST", "SGST", "IGST", "TotalInvoiceAmount"]:
            col_idx = df_comparing_totals.columns.get_loc(each)
            writer.sheets['Ezy Invoicing Summary'].set_column(
                col_idx, col_idx, 20)
        invoice_comparison = compare_invoice_summary(start_date, end_date)
        if not invoice_comparison["success"]:
            return invoice_comparison
        comparing_inv = [{k.replace(' ', '_'): v for k, v in d.items()}
                         for d in invoice_comparison["data"]]
        recon_data["Comparison"] = comparing_inv
        df_invoice_comparison = pd.DataFrame.from_records(
            invoice_comparison["data"])
        df_invoice_comparison = df_invoice_comparison.reindex(columns=["InvoiceNumber", "EzyinvoicingBaseAmount", "OperaBaseAmount", "BaseMissmatchAmount", "BaseAmountStatus", "EzyinvoicingInvoiceAmount",
                                                              "OperaInvoiceAmount", "InvoiceMissmatchAmount", "InvoiceAmountStatus", "IRN Status", "IRN Number", "Acknowledgement No", "Acknowledgement Date"])
        df_invoice_comparison.to_excel(
            writer, sheet_name="Comparison", index=False)
        for each in ["InvoiceNumber", "EzyinvoicingBaseAmount", "OperaBaseAmount", "BaseMissmatchAmount", "BaseAmountStatus", "EzyinvoicingInvoiceAmount", "OperaInvoiceAmount", "InvoiceMissmatchAmount", "InvoiceAmountStatus", "IRN Status", "IRN Number", "Acknowledgement No", "Acknowledgement Date"]:
            col_idx = df_invoice_comparison.columns.get_loc(each)
            writer.sheets['Comparison'].set_column(col_idx, col_idx, 25)
        conb2bb2c = [{k.replace(' ', '_'): v for k, v in d.items()}
                     for d in missing["data"]["converted_b2b_to_b2c"]]
        recon_data["Converted_B2B_to_B2C"] = conb2bb2c
        df_b2b_to_b2c = pd.DataFrame.from_records(
            missing["data"]["converted_b2b_to_b2c"])
        df_b2b_to_b2c.to_excel(
            writer, sheet_name="Converted B2B to B2C", index=False)
        for each in ["InvoiceNumber", "InvoiceUploadType", "PrintedDate", "ConvertedDate"]:
            col_idx = df_b2b_to_b2c.columns.get_loc(each)
            writer.sheets['Converted B2B to B2C'].set_column(
                col_idx, col_idx, 20)
        conb2cb2b = [{k.replace(' ', '_'): v for k, v in d.items()}
                     for d in missing["data"]["converted_b2c_to_b2b"]]
        recon_data["Converted_B2C_to_B2B"] = conb2cb2b
        df_b2c_to_b2b = pd.DataFrame.from_records(
            missing["data"]["converted_b2c_to_b2b"])
        df_b2c_to_b2b.to_excel(
            writer, sheet_name="Converted B2C to B2B", index=False)
        for each in ["InvoiceNumber", "InvoiceUploadType", "PrintedDate", "ConvertedDate"]:
            col_idx = df_b2c_to_b2b.columns.get_loc(each)
            writer.sheets['Converted B2C to B2B'].set_column(
                col_idx, col_idx, 20)
        hsn = getHsnSummary(start_date=start_date, end_date=end_date)
        if not hsn["success"]:
            return hsn
        if len(hsn["data"]) == 0:
            hsn["data"] = [{"SAC Code": "", "Gst Rate": "", "UQC": "", "Total Quantity": "", "CGST Amount": "", "SGST Amount": "", "IGST Amount": "",
                            "State CESS Amount": "", "Central CESS Amount": "", "VAT": "", "Total GST": "", "Total Tax Amount": "", "Total Amount": ""}]
        df_hsn = pd.DataFrame.from_records(hsn["data"])
        total = df_hsn.sum(numeric_only=True, axis=0)
        total.name = "Total"
        hsn_total = []
        if bool(dict(total.transpose())):
            hsn_total = [dict(total.transpose())]
        hsn_summary_data = hsn["data"]+hsn_total
        hsnb2b = [{k.replace(' ', '_'): v for k, v in d.items()}
                  for d in hsn_summary_data]
        recon_data["HSN"] = hsnb2b
        df_b2bc = df_hsn.append(total.transpose())
        if len(hsn["data"]) > 1:
            df_b2bc['SAC Code'] = df_b2bc['SAC Code'].replace(
                np.nan, "Total")
            df_b2bc['Gst Rate'] = np .where(
                df_b2bc['Gst Rate'] > 25, "", df_b2bc["Gst Rate"])
        df_b2bc.to_excel(
            writer, sheet_name="HSN", index=False)
        for each in ["SAC Code", "Gst Rate", "UQC", "Total Quantity", "CGST Amount", "SGST Amount", "IGST Amount", "State CESS Amount", "Central CESS Amount", "VAT", "Total GST", "Total Tax Amount", "Total Amount"]:
            col_idx = df_b2bc.columns.get_loc(each)
            writer.sheets['HSN'].set_column(col_idx, col_idx, 20)
        hsn = getHsnSummary(
            filters=[["invoice_type", "=", "B2B"]],  start_date=start_date, end_date=end_date)
        if not hsn["success"]:
            return hsn
        if len(hsn["data"]) == 0:
            hsn["data"] = [{"SAC Code": "", "Gst Rate": "", "UQC": "", "Total Quantity": "", "CGST Amount": "", "SGST Amount": "", "IGST Amount": "",
                            "State CESS Amount": "", "Central CESS Amount": "", "VAT": "", "Total GST": "", "Total Tax Amount": "", "Total Amount": ""}]
        df_b2b_hsn = pd.DataFrame.from_records(hsn["data"])
        total = df_b2b_hsn.sum(numeric_only=True, axis=0)
        total.name = "Total"
        b2b_hsn_total = []
        if bool(dict(total.transpose())):
            b2b_hsn_total = [dict(total.transpose())]
        b2b_hsn_summary_data = hsn["data"]+b2b_hsn_total
        hsnb2b = [{k.replace(' ', '_'): v for k, v in d.items()}
                  for d in b2b_hsn_summary_data]
        recon_data["B2B HSN"] = hsnb2b
        df_b2b = df_b2b_hsn.append(total.transpose())
        if len(hsn["data"]) > 1:
            df_b2b['SAC Code'] = df_b2b['SAC Code'].replace(
                np.nan, "Total")
            df_b2b['Gst Rate'] = np.where(
                df_b2b['Gst Rate'] > 25, "", df_b2b["Gst Rate"])
        df_b2b.to_excel(
            writer, sheet_name="B2B HSN", index=False)
        for each in ["SAC Code", "Gst Rate", "UQC", "Total Quantity", "CGST Amount", "SGST Amount", "IGST Amount", "State CESS Amount", "Central CESS Amount", "VAT", "Total GST", "Total Tax Amount", "Total Amount"]:
            col_idx = df_b2b.columns.get_loc(each)
            writer.sheets['B2B HSN'].set_column(col_idx, col_idx, 20)
        b2c_hsn = getHsnSummary(
            filters=[["invoice_type", "=", "B2C"]],  start_date=start_date, end_date=end_date)
        if not b2c_hsn["success"]:
            return b2c_hsn
        if len(b2c_hsn["data"]) == 0:
            b2c_hsn["data"] = [{"SAC Code": "", "Gst Rate": "", "UQC": "", "Total Quantity": "", "CGST Amount": "", "SGST Amount": "", "IGST Amount": "",
                                "State CESS Amount": "", "Central CESS Amount": "", "VAT": "", "Total GST": "", "Total Tax Amount": "", "Total Amount": ""}]
        df_b2c_hsn = pd.DataFrame.from_records(b2c_hsn["data"])
        total_b2c = df_b2c_hsn.sum(numeric_only=True, axis=0)
        total_b2c.name = "Total"
        b2c_hsn_total = []
        if bool(dict(total.transpose())):
            b2c_hsn_total = [dict(total.transpose())]
        b2c_hsn_summary_data = b2c_hsn["data"]+b2c_hsn_total
        hsnb2c = [{k.replace(' ', '_'): v for k, v in d.items()}
                  for d in b2c_hsn_summary_data]
        recon_data["B2C_HSN"] = hsnb2c
        df_b2c = df_b2c_hsn.append(total_b2c.transpose())
        if len(b2c_hsn["data"]) > 1:
            df_b2c['SAC Code'] = df_b2c['SAC Code'].replace(
                np.nan, "Total")
            df_b2c['Gst Rate'] = np.where(
                df_b2c['Gst Rate'] > 25, "", df_b2c["Gst Rate"])
        df_b2c.to_excel(
            writer, sheet_name="B2C HSN", index=False)
        for each in ["SAC Code", "Gst Rate", "UQC", "Total Quantity", "CGST Amount", "SGST Amount", "IGST Amount", "State CESS Amount", "Central CESS Amount", "VAT", "Total GST", "Total Tax Amount", "Total Amount"]:
            col_idx = df_b2c.columns.get_loc(each)
            writer.sheets['B2C HSN'].set_column(col_idx, col_idx, 20)
        no_sac_query = frappe.db.sql("""SELECT `tabItems`.parent as InvoiceNumber, `tabItems`.description as ItemDescription, `tabItems`.item_value_after_gst as BeforeGST, `tabItems`.cgst_amount as CGSTAmount, `tabItems`.sgst_amount as SGSTAmount, `tabItems`.igst_amount as IGSTAmount,`tabItems`.item_taxable_value as TaxableValue, `tabInvoices`.irn_generated as IRNstatus from `tabItems` INNER JOIN `tabInvoices` ON `tabItems`.parent = `tabInvoices`.invoice_number where `tabItems`.sac_code = "No Sac" and invoice_date between '{}' and '{}'""".format(start_date, end_date), as_dict=1)

        # print(no_sac_query,"//////.....")
        # invoices_no_sac = frappe.db.get_list('Items', filters=[["date", 'between', [start_date, end_date]],["sac_code","=","No Sac"]], fields=["`tabItems`.parent as InvoiceNumber", "`tabItems`.description as ItemDescription", "`tabItems`.item_value as BeforeGST", "`tabItems`.item_value_after_gst as AfterGST", "`tabInvoices`.irn_generated as IRNstatus"])
        no_sac = [{k.replace(' ', '_'): v for k, v in d.items()}
                  for d in no_sac_query]
        recon_data["No_Sac"] = no_sac
        sac_no = pd.DataFrame.from_dict(no_sac)
        sac_no.to_excel(
            writer, sheet_name="No Sac", index=False)
        for each in ["InvoiceNumber", "ItemDescription", "BeforeGST", "CGSTAmount", "SGSTAmount", "IGSTAmount", "TaxableValue", "IRNstatus"]:
            col_idx = sac_no.columns.get_loc(each)
            writer.sheets['No Sac'].set_column(
                col_idx, col_idx, 20)
        # document = document_sequence(start_date, end_date)
        # document_seq = [{k.replace(' ', '_') : v for k, v in d.items()} for d in document]
        # recon_data["Sequence"] = document_seq
        # document_fields = ["Document","From","To","Success","Cancelled", "Error"]
        # sequence_count = [{"Document":"Tax Invoice", "From":document_sequence["data"]["tax_invoice_from"], "To":document_sequence["data"]["tax_invoice_to"], "Success":document_sequence["data"]["tax_invoice_success_count"], "Cancelled":document_sequence["data"]["tax_invoice_cancelled_count"], "Error":document_sequence["data"]["tax_invoice_error_count"]},{"Document":"Credit Invoice", "From":document_sequence["data"]["credit_invoice_from"], "To":document_sequence["data"]["credit_invoice_to"], "Success":document_sequence["data"]["credit_invoice_success_count"], "Cancelled":document_sequence["data"]["credit_invoice_cancelled_count"], "Error":document_sequence["data"]["credit_invoice_error_count"]}]
        # df_b2c.to_excel(
        #     writer, sheet_name="Sequence", index=False)
        # for each in ["Document","From","To","Success","Cancelled", "Error"]:
        #     col_idx = df_b2c.columns.get_loc(each)
        #     writer.sheets['B2C HSN'].set_column(col_idx, col_idx, 20)
        writer.save()
        if workbook == "true":
            files_new = {"file": open(file_path, 'rb')}
            payload_new = {'is_private': 1, 'folder': 'Home'}
            file_response = requests.post(company.host+"api/method/upload_file", files=files_new,
                                          data=payload_new, verify=False).json()
            if "file_url" in file_response["message"].keys():
                os.remove(file_path)
                return {"success": True, "file_url": file_response["message"]["file_url"], "file_name": "RECON-"+month_name+"-"+year_object+".xlsx"}
            return {"success": False, "message": "something went wrong"}
        else:
            os.remove(file_path)
            return {"success": True, "data": recon_data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("document_sequence",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


# def get_missing_invoices(month=None, year=None):
#     try:
#         data = {}
#         start_date = year+'-'+month+"-01"
#         end_date = date_util.get_last_day(start_date)
#         invoice_filters = {"tax_invoices_in_ezyinvoicing": [["invoice_date", 'between', [start_date, end_date]], [
#             "invoice_category", "=", "Tax Invoice"]], "credit_invoices_in_ezyinvoicing": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "Credit Invoice"]]}
#         for key, value in invoice_filters.items():
#             ezy_invoicing_invoices = frappe.db.get_list(
#                 'Invoices', filters=value, pluck='name', order_by='name')
#             data.update({key: ezy_invoicing_invoices})
#         opera_folios_filters = {"opera_tax_folios": [["bill_generation_date", "between", [start_date, end_date]], [
#             "folio_type", "=", "TAX INVOICE"]], "opera_credit_folios": [["bill_generation_date", "between", [start_date, end_date]], ["folio_type", "=", "CREDIT INVOICE"]]}
#         for key, value in opera_folios_filters.items():
#             opera_folios = frappe.db.get_list(
#                 "Invoice Reconciliations", filters=value, pluck="name", order_by='name')
#             data.update({key: opera_folios})
#         data["type_missmatch_for_tax"] = frappe.db.get_list('Invoices', filters=[["name", "in", data["opera_tax_folios"]], [
#                                                             "invoice_category", "!=", "Tax Invoice"]], fields=['name as InvoiceNumber',"invoice_category as Ezyinvoicing"], order_by='name', as_list=False)
#         data["type_missmatch_for_credit"] = frappe.db.get_list('Invoices', filters=[["name", "in", data["opera_credit_folios"]], [
#                                                                "invoice_category", "!=", "Credit Invoice"]], fields=['name as InvoiceNumber',"invoice_category as Ezyinvoicing"], order_by='name', as_list=False)
#         if len(data["type_missmatch_for_credit"])>0:
#             data["type_missmatch_for_credit"] = [dict(item, **{'Opera':'Credit Invoice'}) for item in data["type_missmatch_for_credit"]]
#         else:
#             data["type_missmatch_for_credit"] = [{"InvoiceNumber":"","Ezyinvoicing":"","Opera":""}]
#         if len(data["type_missmatch_for_tax"])>0:
#             data["type_missmatch_for_tax"] = [dict(item, **{'Opera':'Tax Invoice'}) for item in data["type_missmatch_for_tax"]]
#         data["missing_count_tax_invoices"] = list(
#             set(data["tax_invoices_in_ezyinvoicing"]) ^ set(data["opera_tax_folios"]))
#         data["missing_count_credit_invoices"] = list(
#             set(data["credit_invoices_in_ezyinvoicing"]) ^ set(data["opera_credit_folios"]))
#         data["missing_tax_opera"] = list(
#             set(data["tax_invoices_in_ezyinvoicing"]) - set(data["opera_tax_folios"]))
#         data["missing_credit_opera"] = list(
#             set(data["credit_invoices_in_ezyinvoicing"]) - set(data["opera_credit_folios"]))
#         data["missing_tax_ezyinvoicing"] = list(
#             set(data["opera_tax_folios"]) - set(data["tax_invoices_in_ezyinvoicing"]))
#         data["missing_credit_ezyinvoicing"] = list(
#             set(data["opera_credit_folios"]) - set(data["credit_invoices_in_ezyinvoicing"]))
#         data["converted_b2b_to_b2c"] = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]],["converted_from_b2b","=","Yes"]], fields=['name as InvoiceNumber',"invoice_from as InvoiceUploadType"], order_by='name')
#         if len(data["converted_b2b_to_b2c"])==0:
#             data["converted_b2b_to_b2c"] = [{"InvoiceNumber":"", "InvoiceUploadType":""}]
#         data["converted_b2c_to_b2b"] = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]],["converted_from_b2c","=","Yes"]], fields=['name as InvoiceNumber',"invoice_from as InvoiceUploadType"], order_by='name')
#         if len(data["converted_b2c_to_b2b"])==0:
#             data["converted_b2c_to_b2b"] = [{"InvoiceNumber":"", "InvoiceUploadType":""}]
#         return {"success": True, "data": data}
#     except Exception as e:
#         exc_type, exc_obj, exc_tb = sys.exc_info()
#         frappe.log_error("get_missing_invoices",
#                          "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
#         return {"success": False, "message": str(e)}

def get_missing_invoices(start_date, end_date):
    try:
        data = {}
        invoice_filters = {"tax_invoices_in_ezyinvoicing": [["invoice_date", 'between', [start_date, end_date]], [
            "invoice_category", "=", "Tax Invoice"]], "credit_invoices_in_ezyinvoicing": [["invoice_date", 'between', [start_date, end_date]], ["invoice_category", "=", "Credit Invoice"]]}
        for key, value in invoice_filters.items():
            ezy_invoicing_invoices = frappe.db.get_list(
                'Invoices', filters=value, pluck='name', order_by='name')
            data.update({key: ezy_invoicing_invoices})
        opera_folios_filters = {"opera_tax_folios": [["bill_generation_date", "between", [start_date, end_date]], [
            "folio_type", "=", "TAX INVOICE"]], "opera_credit_folios": [["bill_generation_date", "between", [start_date, end_date]], ["folio_type", "=", "CREDIT INVOICE"]]}
        for key, value in opera_folios_filters.items():
            opera_folios = frappe.db.get_list(
                "Invoice Reconciliations", filters=value, pluck="name", order_by='name')
            data.update({key: opera_folios})
        data["opera_folios"] = frappe.db.get_list("Invoice Reconciliations", filters=[
                                                  ["bill_generation_date", "between", [start_date, end_date]]], pluck="name", order_by='name')
        data["ezy_invoicing_invoices"] = frappe.db.get_list('Invoices', filters=[
                                                            ["invoice_date", 'between', [start_date, end_date]]], pluck='name', order_by='name')
        data["type_missmatch_for_tax"] = frappe.db.get_list('Invoices', filters=[["name", "in", data["opera_tax_folios"]], [
                                                            "invoice_category", "!=", "Tax Invoice"]], fields=['name as invoicenumber', "invoice_category as ezyinvoicing"], order_by='name', as_list=False)
        data["type_missmatch_for_credit"] = frappe.db.get_list('Invoices', filters=[["name", "in", data["opera_credit_folios"]], [
                                                               "invoice_category", "!=", "Credit Invoice"]], fields=['name as invoicenumber', "invoice_category as ezyinvoicing"], order_by='name', as_list=False)
        if len(data["type_missmatch_for_credit"]) > 0:
            data["type_missmatch_for_credit"] = [dict(
                item, **{'opera': 'Credit Invoice'}) for item in data["type_missmatch_for_credit"]]
        else:
            data["type_missmatch_for_credit"] = [
                {"InvoiceNumber": "", "Opera": "", "Ezyinvoicing": ""}]
        if len(data["type_missmatch_for_tax"]) > 0:
            data["type_missmatch_for_tax"] = [dict(
                item, **{'opera': 'Tax Invoice'}) for item in data["type_missmatch_for_tax"]]
        data["missing_in_opera"] = list(
            set(data["ezy_invoicing_invoices"]) - set(data["opera_folios"]))
        data["missing_in_ezyinvoicing"] = list(
            set(data["opera_folios"]) - set(data["ezy_invoicing_invoices"]))
        data["converted_b2b_to_b2c"] = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]], [
                                                          "converted_from_b2b", "=", "Yes"]], fields=['name as invoicenumber', "invoice_from as invoiceuploadtype", "ack_date as printeddate", "converted_from_b2b_time as converteddate"], order_by='name')
        if len(data["converted_b2b_to_b2c"]) == 0:
            data["converted_b2b_to_b2c"] = [
                {"InvoiceNumber": "", "InvoiceUploadType": "", "PrintedDate": "", "ConvertedDate": ""}]
        data["converted_b2c_to_b2b"] = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]], [
                                                          "converted_from_b2c", "=", "Yes"]], fields=['name as invoicenumber', "invoice_from as invoiceuploadtype", "ack_date as printeddate", "converted_from_b2c_time as converteddate"], order_by='name')
        if len(data["converted_b2c_to_b2b"]) == 0:
            data["converted_b2c_to_b2b"] = [
                {"InvoiceNumber": "", "InvoiceUploadType": "", "PrintedDate": "", "ConvertedDate": ""}]
        return {"success": True, "data": data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("get_missing_invoices",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


def invoices_summary(start_date, end_date):
    try:
        ezy_invoicing_invoices = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]]], fields=[
                                                    "sum(sales_amount_before_tax) as BeforeTax", "sum(cgst_amount) as CGST", "sum(sgst_amount) as SGST", "sum(igst_amount) as IGST", "sum(sales_amount_after_tax) as TotalInvoiceAmount"], order_by='name')
        return {"success": True, "data": ezy_invoicing_invoices}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("invoices_summary",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


def compare_invoice_summary(start_date, end_date):
    try:
        get_invoices = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]]], fields=[
                                          "name as invoicenumber", "(sales_amount_before_tax+total_central_cess_amount+total_state_cess_amount+total_credit_vat_amount) as ezyinvoicingbaseamount", "sales_amount_after_tax as ezyinvoicinginvoiceamount"], order_by='name')
        get_payments = frappe.db.get_list("Payment Types", pluck="name")
        if len(get_invoices) > 0:
            for each in get_invoices:
                get_folios = frappe.db.get_value("Invoice Reconciliations", {"name": each["invoicenumber"]}, [
                                                 "name as invoicenumber", "total_invoice_amount"], order_by='name', as_dict=True)
                if get_folios:
                    each["operainvoiceamount"] = get_folios["total_invoice_amount"]
                    each["invoicemissmatchamount"] = round(
                        each["ezyinvoicinginvoiceamount"]-get_folios["total_invoice_amount"], 2)
                    if each["invoicemissmatchamount"] > -1 and each["invoicemissmatchamount"] < 1:
                        each["invoiceamountstatus"] = "Match"
                    else:
                        each["invoiceamountstatus"] = "MisMatch"
                    get_items = frappe.db.get_list("Opera Invoice Items", filters=[["invoice_reconciliations", "=", each["invoicenumber"]], [
                                                   "transaction_description", "not in", get_payments]], fields=["sum(debit_amount) as BaseAmount"])
                    if len(get_items) > 0:
                        each["operabaseamount"] = get_items[0]["BaseAmount"] if get_items[0]["BaseAmount"] else 0
                        each["basemissmatchamount"] = round(
                            each["ezyinvoicingbaseamount"]-each["operabaseamount"], 2)
                        if each["basemissmatchamount"] > -1 and each["basemissmatchamount"] < 1:
                            each["baseamountstatus"] = "Match"
                        else:
                            each["baseamountstatus"] = "MisMatch"
                    else:
                        each["operabaseamount"] = 0
                    # each["BaseAmountStatus"] = ""
                else:
                    each.update({"operainvoiceamount": 0, "invoicemissmatchamount": 0, "invoiceamountstatus": "Missing in Opera",
                                "operabaseamount": 0, "basemissmatchamount": 0, "baseamountstatus": "Missing in Opera"})
                invoice_columns = frappe.db.sql(
                    """select irn_generated as 'irn_status',If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as 'irn_number',If(invoice_category!="Credit Invoice",ack_no,credit_ack_no) as 'acknowledgement_no',  If(invoice_category!="Credit Invoice",ack_date,credit_ack_date) as 'acknowledgement_date' from `tabInvoices` where name='{}' ORDER BY name""".format(each["invoicenumber"]), as_dict=1)
                if len(invoice_columns) > 0:
                    each.update(invoice_columns[0])
                else:
                    each.update({"irn_status": "", "irn_number": "",
                                "acknowledgement_no": "", "acknowledgement_date": ""})
            return {"success": True, "data": get_invoices}
        else:
            data = [{"invoicenumber": "", "ezyinvoicingbaseamount": "", "ezyinvoicinginvoiceamount": "", "operainvoiceamount": "", "invoicemissmatchamount": "",
                     "invoiceamountstatus": "", "operabaseamount": "", "basemissmatchamount": "", "baseamountstatus": ""}]
            return {"success": True, "data": data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("compare_invoice_summary",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


# @frappe.whitelist(allow_guest=True)
# def count(start_date, end_date, export=False):
#     try:
#         company = frappe.get_last_doc("company")
#         cwd = os.getcwd()
#         site_name = cstr(frappe.local.site)
#         month_name = datetime.datetime.strptime(
#             start_date, "%Y-%m-%d").strftime("%b")
#         year_object = datetime.datetime.strptime(
#             start_date, "%Y-%m-%d").strftime("%Y")
#         file_path = cwd + "/" + site_name + \
#             "/public/files/Count-"+month_name+"-"+year_object+".xlsx"
#         missing = get_missing_invoices(start_date, end_date)
#         if not missing["success"]:
#             return missing
#         # return missing
#         recon_data = {}
#         if export == False:
#             sequence_data = [{"Ezyinvoicing Count": len(missing["data"]["ezy_invoicing_invoices"]), "Opera Folios Count": len(missing["data"]["opera_folios"]), "Missing In EzyInvoicing": len(missing["data"]["missing_in_ezyinvoicing"]), "Missing In Opera": len(
#             missing["data"]["missing_in_opera"]), "Type Missmatch": len(missing["data"]["type_missmatch_for_tax"])+len(missing["data"]["type_missmatch_for_credit"]) if len(missing["data"]["type_missmatch_for_credit"]) > 1 else 0}]
#         else:
#         # Summary of Invoices
#         print(sequence_data,".........///////.....")
#         writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
#         seq_data = [{k.replace(' ', '_') : v for k, v in d.items()} for d in sequence_data]
#         recon_data["count"] = seq_data
#         df_summary = pd.DataFrame(sequence_data)
#         print(df_summary,";;;;;;;;;;;:::::::::;")
#         df_summary.to_excel(writer, sheet_name="Count", index=False)
#         for each in ["Ezyinvoicing Count", "Opera Folios Count", "Missing In EzyInvoicing", "Missing In Opera", "Type Missmatch"]:
#             col_idx = df_summary.columns.get_loc(each)
#             writer.sheets['Count'].set_column(col_idx, col_idx, 20)

#     except Exception as e:
#         exc_type, exc_obj, exc_tb = sys.exc_info()
#         frappe.log_error("compare_invoice_summary",
#                          "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
#         return {"success": False, "message": str(e)}

@frappe.whitelist(allow_guest=True)
def create_reconciliation(start_date=None, end_date=None, redo=False, recon_id=None):
    try:
        # Getting recon ID
        if (start_date and end_date) or (redo and recon_id):
            enqueue(
                    create_recon,
                    queue="default",
                    timeout=800000,
                    event="create_recon",
                    now=False,
                    data={"start_date": start_date, "end_date": end_date, "redo": False, "recon_id": recon_id},
                    is_async=True,
                )
            return {"success": True, "message": "Reconciliation created"}
        return {"success": False, "message": "something went wrong"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("document_sequence",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}
    
def create_recon(data):
    try:
        start_date=data["start_date"]
        end_date=data["end_date"]
        redo=data["redo"]
        recon_id=data["recon_id"]
        if not redo:
            insert_recon_data = insert_records(
                {"doctype": "Recon Details", "start_date": start_date, "end_date": end_date, "user_id": frappe.session.user})
            if not insert_recon_data["success"]:
                return insert_recon_data
            frappe.publish_realtime("custom_socket", {"message": "Recon Processing"})
            recon_doc = insert_recon_data["doc"]
        else:
            if not recon_id:
                return {"success": True, "message": "reconID is mandatory"}
            recon_doc = frappe.get_doc("Recon Details", recon_id)
            start_date = recon_doc.start_date
            end_date = recon_doc.end_date
        missing = get_missing_invoices(start_date, end_date)
        if not missing["success"]:
            return missing

        # Inserting Recon ID into Invoice Count
        invoice_count_data = {"doctype": "Invoice Count", "ezyinvoicing_count": len(missing["data"]["ezy_invoicing_invoices"]), "opera_folios_count": len(missing["data"]["opera_folios"]), "missing_in_ezyinvoicing": len(missing["data"]["missing_in_ezyinvoicing"]), "missing_in_opera": len(
            missing["data"]["missing_in_opera"]), "type_missmatch": len(missing["data"]["type_missmatch_for_tax"])+len(missing["data"]["type_missmatch_for_credit"]) if len(missing["data"]["type_missmatch_for_credit"]) > 1 else 0, "recon_id": recon_doc.name}
        insert_invoice_count = insert_records(invoice_count_data)
        if not insert_invoice_count["success"]:
            return insert_invoice_count

        # Missing Invoices In Opera
        # missing_in_opera = {"invoices_numbers": missing["data"]["missing_in_opera"]}
        if len(missing["data"]["missing_in_opera"]) > 0:
            missing_numbers = tuple(missing["data"]["missing_in_opera"])
            if len(missing_numbers) == 1:
                missing_numbers = missing_numbers + ('0',)
            get_missing_in_opera = frappe.db.sql("""select invoice_number, invoice_date,invoice_type, invoice_category, irn_generated as 'irn_status', invoice_from as 'ezyinvoice_status', pms_invoice_summary as 'invoice_amount', pms_invoice_summary_without_gst as 'taxable_value', igst_amount, sgst_amount, cgst_amount, total_gst_amount, other_charges, (total_central_cess_amount+total_state_cess_amount) as cess,total_vat_amount as 'vat_amount', If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as 'irn_number',If(invoice_category!="Credit Invoice",ack_no,credit_ack_no) as 'acknowledgement_no',  If(invoice_category!="Credit Invoice",ack_date,credit_ack_date) as 'acknowledgement_date' from `tabInvoices` where name in {} and irn_generated != 'Zero Invoice' ORDER BY name""".format(missing_numbers), as_dict=1)
            if len(get_missing_in_opera) > 0:
                for each in get_missing_in_opera:
                    each["doctype"] = "Missing in Opera"
                    each["recon_id"] = recon_doc.name
                    missing_in_opera = insert_records(each)
                    if not missing_in_opera["success"]:
                        return missing_in_opera

        # Zero Invoices
        missing_in_opera = {
            "Invoices Numbers": missing["data"]["missing_in_opera"]}
        get_missing_in_opera_data = [{"Invoice Number": "", "Invoice Date": "", "Invoice Type": "", "Invoice Category": "", "IRN Status": "", "EzyInvoice Status": "", "Invoice Amount": "", "Taxable Value": "",
                                      "IGST Amount": "", "SGST Amount": "", "CGST Amount": "", "Total Gst Amount": "", "Other Charges": "", "CESS": "", "VAT Amount": "", "IRN Number": "", "Acknowledgement No": "", "Acknowledgement Date": ""}]
        if len(missing["data"]["missing_in_opera"]) > 0:
            missing_numbers = tuple(missing["data"]["missing_in_opera"])
            if len(missing_numbers) == 1:
                missing_numbers = missing_numbers + ('0',)
            get_missing_in_opera = frappe.db.sql("""select invoice_number, invoice_date,invoice_type, invoice_category, irn_generated as 'irn_status', invoice_from as 'ezyinvoice_status', pms_invoice_summary as 'invoice_amount', pms_invoice_summary_without_gst as 'taxable_value', igst_amount, sgst_amount, cgst_amount, total_gst_amount, other_charges, (total_central_cess_amount+total_state_cess_amount) as cess,total_vat_amount as 'vat_amount', If(invoice_category!="Credit Invoice",irn_number,credit_irn_number) as 'irn_number',If(invoice_category!="Credit Invoice",ack_no,credit_ack_no) as 'acknowledgement_no',  If(invoice_category!="Credit Invoice",ack_date,credit_ack_date) as 'acknowledgement_date' from `tabInvoices` where name in {} and irn_generated = 'Zero Invoice' ORDER BY name""".format(missing_numbers), as_dict=1)
            if len(get_missing_in_opera) > 0:
                for each in get_missing_in_opera:
                    each["doctype"] = "Missing in Opera"
                    each["recon_id"] = recon_doc.name
                    each["is_zero_invoice"] = 1
                    zero_invoices = insert_records(each)
                    print(zero_invoices, "////////")
                    if not zero_invoices["success"]:
                        return zero_invoices

        # Missing Invoices In Ezyinvoicing
        if len(missing["data"]["missing_in_ezyinvoicing"]) > 0:
            for each in missing["data"]["missing_in_ezyinvoicing"]:
                missing_in_ezy = {"doctype": "Missing in EzyInvoicing",
                                  "invoice_number": each, "recon_id": recon_doc.name}
                missing_in_ezyinvoicing = insert_records(missing_in_ezy)
                if not missing_in_ezyinvoicing["success"]:
                    return missing_in_ezyinvoicing

        # Type Missmatch
        type_missmatch = missing["data"]["type_missmatch_for_tax"] + \
            missing["data"]["type_missmatch_for_credit"]
        if len(type_missmatch) > 0:
            for each in type_missmatch:
                each["doctype"] = "Invoice Type Missmatch"
                each["recon_id"] = recon_doc.name
                type_mismatch = insert_records(each)
                if not type_mismatch["success"]:
                    return type_mismatch

        # Ezy Invoicing Summary
        ezy_invoicing_invoices = frappe.db.get_list('Invoices', filters=[["invoice_date", 'between', [start_date, end_date]]], fields=[
                                                    "sum(sales_amount_before_tax) as beforetax", "sum(cgst_amount) as sgst", "sum(sgst_amount) as cgst", "sum(igst_amount) as igst", "sum(sales_amount_after_tax) as totalinvoiceamount"], order_by='name')
        if len(ezy_invoicing_invoices) > 0:
            for each in ezy_invoicing_invoices:
                each["doctype"] = "Ezy Invoicing Summary"
                each["recon_id"] = recon_doc.name
                compare = insert_records(each)
                if not compare["success"]:
                    return compare

        # Comparison
        invoice_comparison = compare_invoice_summary(start_date, end_date)
        if len(invoice_comparison["data"]) > 0:
            for each in invoice_comparison["data"]:
                each["doctype"] = "Recon Opera Comparison"
                each["recon_id"] = recon_doc.name
                comparsion = insert_records(each)
                if not comparsion["success"]:
                    return comparsion

        # Converted b2b to b2c
        if len(missing["data"]["converted_b2b_to_b2c"]) > 0:
            for each in missing["data"]["converted_b2b_to_b2c"]:
                each["doctype"] = "Converted B2B to B2C"
                each["recon_id"] = recon_doc.name
                conb2b2c = insert_records(each)
                if not conb2b2c["success"]:
                    return conb2b2c

        # Converted b2c to b2b
        if len(missing["data"]["converted_b2c_to_b2b"]) > 0:
            for each in missing["data"]["converted_b2c_to_b2b"]:
                each["doctype"] = "Converted B2C to B2B"
                each["recon_id"] = recon_doc.name
                conb2cb2b = insert_records(each)
                if not conb2cb2b["success"]:
                    return conb2cb2b

        # HSN SUmmary
        hsn = getHsnSummary(start_date=start_date, end_date=end_date)
        if len(hsn["data"]) > 0:
            for each in hsn["data"]:
                print(each, ".......................;;;;;;;;;;;")
                each["doctype"] = "Recon HSN Summary"
                each["recon_id"] = recon_doc.name
                each["type"] = "ALL"
                hsnsum = insert_records(each)
                if not hsnsum["success"]:
                    return hsnsum

         # B2B HSN Summary
        hsnb2b = getHsnSummary(
            filters=[["invoice_type", "=", "B2B"]],  start_date=start_date, end_date=end_date)
        if len(hsnb2b["data"]) > 0:
            for each in hsnb2b["data"]:
                print(each, ".......................;;;;;;;;;;;")
                each["doctype"] = "Recon HSN Summary"
                each["recon_id"] = recon_doc.name
                each["type"] = "B2B"
                hsnsumb2b = insert_records(each)
                if not hsnsumb2b["success"]:
                    return hsnsumb2b

        # B2C HSN Summary
        hsnb2c = getHsnSummary(
            filters=[["invoice_type", "=", "B2C"]],  start_date=start_date, end_date=end_date)
        if len(hsnb2c["data"]) > 0:
            for each in hsnb2c["data"]:
                print(each, ".......................;;;;;;;;;;;")
                each["doctype"] = "Recon HSN Summary"
                each["recon_id"] = recon_doc.name
                each["type"] = "B2C"
                hsnsumb2c = insert_records(each)
                if not hsnsumb2c["success"]:
                    return hsnsumb2c

        # No SAC
        no_sac_query = frappe.db.sql("""SELECT `tabItems`.parent as invoicenumber, `tabItems`.description as itemdescription, `tabItems`.item_value_after_gst as before_gst, `tabItems`.cgst_amount as cgst_amount, `tabItems`.sgst_amount as sgst_amount, `tabItems`.igst_amount as igst_amount,`tabItems`.item_taxable_value as total_tax_amount, `tabItems`.vat_amount as vat_amount, `tabInvoices`.irn_generated as irnstatus from `tabItems` INNER JOIN `tabInvoices` ON `tabItems`.parent = `tabInvoices`.invoice_number where `tabItems`.sac_code = "No Sac" and invoice_date between '{}' and '{}'""".format(start_date, end_date), as_dict=1)
        print("......;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;..")
        if len(no_sac_query) > 0:
            for each in no_sac_query:
                each["doctype"] = "No Sac"
                each["recon_id"] = recon_doc.name
                no_sac = insert_records(each)
                if not no_sac["success"]:
                    return no_sac
        frappe.publish_realtime("custom_socket", {"message": "Create Recon"})
        return {'success': True, "message": "created recon"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("create_recon",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


def insert_records(data):
    try:
        doc = frappe.get_doc(data)
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return {"success": True, "doc": doc}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("insert_records",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def export_recon(doctype=None, reconID=None, export="False", filter={}):
    try:
        doc_name = doctype
        if isinstance(filter, str):
            filter = json.loads(filter)
        # filter.extend(["invoice_type", "=", "B2C"])
        filters = {"recon_id": reconID}
        filters.update(filter)
        if doctype == "Invoice Count":
            fields = ["ezyinvoicing_count", "opera_folios_count",
                      "missing_in_ezyinvoicing", "missing_in_opera", "type_missmatch"]
            fields1 = ["ezyinvoicing_count as 'Ezyinvoicing Count'", "opera_folios_count as 'Opera Folio Count'",
                       "missing_in_ezyinvoicing as 'Missing in Ezyinvoicing'", "missing_in_opera as 'Missing in Opera'", "type_missmatch as 'Type Missmatch'"]
        elif doctype in ["Missing in Opera", "Zero Invoice"]:
            filters["is_zero_invoice"] = 1 if doctype == "Zero Invoice" else 0
            fields = ["invoice_number", "invoice_date", "invoice_type", "invoice_category", "irn_status", "ezyinvoice_status", "invoice_amount", "total_tax_amount", "igst_amount",
                      "sgst_amount", "cgst_amount", "total_gst_amount", "other_charges", "cess", "vat_amount", "irn_number", "acknowledgement_no", "acknowledgement_date"]
            fields1 = ["invoice_number as 'Invoice Number'", "invoice_date as 'Invoice Date'", "invoice_type as 'Invoice Type'", "invoice_category as 'Invoice Category'", "irn_status as 'IRN Status'", "ezyinvoice_status as 'Ezyinvoice Status'", "invoice_amount as 'Invoice Amount'", "total_tax_amount as 'Taxable Amount'", "igst_amount as 'IGST Amount'",
                       "sgst_amount as 'SGST Amount'", "cgst_amount as 'CGST Amount'", "total_gst_amount as 'Total GST Amount'", "other_charges as 'Other Charges'", "cess as 'CESS'", "vat_amount as 'VAT Amount'", "irn_number as 'IRN Number'", "acknowledgement_no as 'Acknowledge Number'", "acknowledgement_date as 'Acknowledge Date'"]
            doctype = "Missing in Opera"
        elif doctype == "Missing in EzyInvoicing":
            fields = ["invoice_number"]
            fields1 = ["invoice_number as 'Invoice Number'"]
        elif doctype == "Invoice Type Missmatch":
            fields = ["invoicenumber", "ezyinvoicing", "opera"]
            fields1 = ["invoicenumber as 'Invoice Number'",
                       "ezyinvoicing as 'Ezyinvoicing'", "opera as 'Opera'"]
        elif doctype == "Ezy Invoicing Summary":
            fields = ["beforetax", "cgst", "sgst",
                      "igst", "totalinvoiceamount"]
            fields1 = ["beforetax as 'Before Tax'", "cgst as 'CGST Amount'", "sgst as 'SGST Amount'",
                       "igst as 'IGST Amount'", "totalinvoiceamount as 'Total Invoice Amount'"]
        elif doctype == "Recon Opera Comparison":
            fields = ["invoicenumber", "ezyinvoicingbaseamount", "operabaseamount", "basemissmatchamount", "baseamountstatus", "ezyinvoicinginvoiceamount",
                      "operainvoiceamount", "invoicemissmatchamount", "invoiceamountstatus", "irn_status", "irn_number", "acknowledgement_no", "acknowledgement_date"]
            fields1 = ["invoicenumber as 'Invoice Number'", "ezyinvoicingbaseamount as 'Ezyinvoicing Base Amount'", "operabaseamount as 'Opera Base Amount'", "basemissmatchamount as 'Base Amount Difference'", "baseamountstatus as 'Base Amount Status'", "ezyinvoicinginvoiceamount as 'Ezyinvoicing Invoice Amount'",
                       "operainvoiceamount as 'Opera Invoice Amount'", "invoicemissmatchamount as 'Invoice Amount Difference'", "invoiceamountstatus as 'Invoice Amount Status'", "irn_status as 'IRN Status'", "irn_number as 'IRN Number'", "acknowledgement_no as 'Acknowledge Number'", "acknowledgement_date as 'Acknowledge Date'"]
        elif doctype == "Converted B2B to B2C":
            fields = ["invoicenumber", "invoiceuploadtype",
                      "printeddate", "converteddate"]
            fields1 = ["invoicenumber as 'Invoice Number'", "invoiceuploadtype as 'Invoice Upload Type'",
                       "printeddate as 'Printed Date'", "converteddate as 'Converted Date'"]
        elif doctype == "Converted B2C to B2B":
            fields = ["invoicenumber", "invoiceuploadtype",
                      "printeddate", "converteddate"]
            fields1 = ["invoicenumber as 'Invoice Number'", "invoiceuploadtype as 'Invoice Upload Type'",
                       "printeddate as 'Printed Date'", "converteddate as 'Converted Date'"]
        elif doctype in ["Recon HSN Summary", "B2B HSN Summary", "B2C HSN Summary"]:
            filters["type"] = "ALL" if doctype == "Recon HSN Summary" else (
                "B2B" if doctype == "B2B HSN Summary" else "B2C")
            fields = ["sac_code", "gst_rate", "uqc", "total_quantity", "cgst_amount", "sgst_amount", "igst_amount",
                      "state_cess_amount", "central_cess_amount", "vat_amount", "total_gst_amount", "total_tax_amount", "total_invoice_amount"]
            fields1 = ["sac_code as 'Sac Code'", "gst_rate 'GST Rate'", "uqc as 'UQC'", "total_quantity as 'Total Quantity'", "cgst_amount as 'CGST Amount'", "sgst_amount as 'SGST Amount'", "igst_amount as 'IGST Amount'",
                       "state_cess_amount as 'State Cess Amount'", "central_cess_amount as 'Central Cess Amount'", "vat_amount as 'VAT Amount'", "total_gst_amount as 'Total GST Amount'", "total_tax_amount as 'Total Tax Amount'", "total_invoice_amount as 'Total Invoice Amount'"]
            doctype = "Recon HSN Summary"
        elif doctype == "No Sac":
            fields = ["invoicenumber", "itemdescription", "before_gst",
                      "cgst_amount", "sgst_amount", "igst_amount", "total_tax_amount", "vat_amount", "irnstatus"]
            fields1 = ["invoicenumber as 'Invoice Number'", "itemdescription as 'Item Description'", "before_gst as 'Before GST'", "cgst_amount as 'CGST Amount'",
                       "sgst_amount as 'SGST Amount'", "igst_amount as 'IGST Amount'", "total_tax_amount as 'Taxable Value'", "vat_amount as 'VAT Amount'", "irnstatus as 'IRN Status'"]
        else:
            return {"success": False, "message": "something went wrong"}
        # fields.append("recon_id")
        if export == "True":
            doctype = "tab"+doctype
            sql_filters = " " + \
                (' and '.join("{} {} '{}'".format(key, "=", value)
                 for key, value in filters.items()))
            get_list = frappe.db.sql("""select {} from `{}` where {}""".format(
                ", ".join(fields1), doctype, sql_filters), as_dict=1)
            if len(get_list) > 0:
                company = frappe.get_last_doc("company")
                cwd = os.getcwd()
                site_name = cstr(frappe.local.site)
                file_path = cwd + "/" + site_name + "/public/files/recon_export.xlsx"
                df = pd.DataFrame.from_records(get_list)
                
                writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
                total = df.sum(numeric_only=True, axis=0)
                total.name = "Total"
                df = df.append(total.transpose())
                field_list = list(get_list[0].keys())
                df[field_list[0]] = df[field_list[0]].replace(np.nan, "Total")
                df.to_excel(writer, sheet_name=doc_name, index=False)

                for column in df:
                    column_length = max(df[column].astype(
                        str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    writer.sheets[doc_name].set_column(
                        col_idx, col_idx, column_length)
                writer.save()
                files_new = {"file": open(file_path, 'rb')}
                payload_new = {'is_private': 1, 'folder': 'Home'}
                file_response = requests.post(company.host+"api/method/upload_file", files=files_new,
                                              data=payload_new, verify=False).json()
                if "file_url" in file_response["message"].keys():
                    os.remove(file_path)
                    return {"success": True, "file_url": file_response["message"]["file_url"], "file_name": reconID+doc_name+".xlsx"}
                return {"success": False, "message": "something went wrong"}
            else:
                return {"success": False, "message": "no data found"}
        else:
            get_list = frappe.db.get_list(
                doctype, filters=filters, fields=fields)
            counts = {"count":len(get_list)}
            if doctype in ["Missing in Opera", "Zero Invoice", "Recon Opera Comparison", "Recon HSN Summary", "B2B HSN Summary", "B2C HSN Summary", "No Sac"]:
                df = pd.DataFrame.from_records(get_list)
                total = df.sum(numeric_only=True, axis=0)
                total.name = "Total"
                df = df.append(total.transpose())
                h_total = []
                if bool(dict(total.transpose())):
                    h_total = [dict(total.transpose())]
                    get_list.extend(h_total)
                if doctype == "Recon Opera Comparison":
                    counts_base = df['baseamountstatus'].value_counts()
                    # counts_invoice = df['invoiceamountstatus'].value_counts()
                    # counts_invoice.rename(columns = {'Match':'Inv_Match','MisMatch':'Inv_MisMatch','Missing in Opera':'Missing_in_Opera'}, inplace = True)
                    # print(counts_invoice,"..............")
                    count_base = counts_base.to_dict()
                    counts.update(count_base)
                    # counts_invoice = df['invoiceamountstatus'].value_counts()
            if len(get_list) > 0:
                return {"success": True, "data": get_list, "fields": fields, "counts": counts}
            else:
                get_list = [{i: "" for i in fields}]
                return {"success": True, "data": get_list, "fields": fields, "counts": {"count":0}}
        return {"success": False, "message": "no data found"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("export_recon",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def export_workbook_recon(reconID=None):
    try:
        doctypes = ["Invoice Count", "Missing in Opera", "Zero Invoice", "Missing in EzyInvoicing", "Invoice Type Missmatch", "Ezy Invoicing Summary",
                    "Recon Opera Comparison", "Converted B2B to B2C", "Converted B2C to B2B", "Recon HSN Summary", "B2B HSN Summary", "B2C HSN Summary", "No Sac"]
        wb = Workbook()
        ws = wb.active
        Sheet = wb['Sheet']
        wb.remove(Sheet)
        for each in doctypes:
            data = export_recon(doctype=each, reconID=reconID, export="False")
            if not data["success"]:
                return data
            if each in ["Missing in Opera", "Zero Invoice", "Recon Opera Comparison", "Recon HSN Summary", "B2B HSN Summary", "B2C HSN Summary", "No Sac"]:
                recon_data = pd.DataFrame.from_records(data["data"])
                if len(recon_data) > 0:
                    recon_data[data["fields"][0]] = recon_data[data["fields"][0]].replace(
                        np.nan, "Total")
                    if each in ["Recon HSN Summary", "B2B HSN Summary", "B2C HSN Summary"]:
                        recon_data['gst_rate'] = np .where(
                            recon_data['gst_rate'] > 25, "", recon_data["gst_rate"])
                        recon_data['total_quantity'] = np .where(
                            recon_data['total_quantity'] > 1, "", recon_data["total_quantity"])
                    data["data"] = recon_data.to_dict('records')
            # export_recon.doctypes()
            ws = wb.create_sheet(each)
            ws.title = each
            for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
                ws.column_dimensions[i].width = 20
            fields = [(each.replace("_", " ")).title()
                      for each in data["fields"]]
            ws.append(fields)
            for product in data["data"]:
                values = (product[k] for k in data["fields"])
                ws.append(values)
        cwd = os.getcwd()
        site_name = cstr(frappe.local.site)
        file_path = cwd + "/" + site_name + \
            "/private/files/RECON-"+reconID+".xlsx"
        wb.save(file_path)
        files_new = {"file": open(file_path, 'rb')}
        payload_new = {'is_private': 1, 'folder': 'Home'}
        company = frappe.get_last_doc("company")
        file_response = requests.post(company.host+"api/method/upload_file", files=files_new,
                                      data=payload_new, verify=False).json()
        if "file_url" in file_response["message"].keys():
            os.remove(file_path)
            return {"success": True, "file_url": file_response["message"]["file_url"], "file_name": reconID+".xlsx"}
        return {"success": False, "message": "something went wrong"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("export_workbook_recon",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def delete_recon_file(name=[]):
    try:
        if not isinstance(name, list):
            name = json.loads(name)
        if len(name) > 0:
            enqueue(
                delete_recon,
                queue="default",
                timeout=800000,
                event="data_import",
                now=False,
                data={"recon_ids": name, "redo": False},
                is_async=True,
            )
            return {"success": True, "message": "Recon Deleted"}
        return {"success": False, "message": "reconID's is mandatory"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("delete_recon_file",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


def delete_recon(data):
    try:
        for ids in data["recon_ids"]:
            if frappe.db.exists("Recon Details", ids):
                doctypes = ["Invoice Count", "Missing in Opera", "Missing in EzyInvoicing", "Invoice Type Missmatch", "Ezy Invoicing Summary",
                            "Recon Opera Comparison", "Converted B2B to B2C", "Converted B2C to B2B", "Recon HSN Summary", "No Sac"]
                for each in doctypes:
                    frappe.db.delete(each, {"recon_id": ids})
                if not data["redo"]:
                    frappe.db.delete("Recon Details", ids)
                    frappe.db.commit()
        frappe.publish_realtime("custom_socket", {"message": "Delete Recon"})
        return {"success": True, "message": "recon deleted"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("delete_recon",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


@frappe.whitelist(allow_guest=True)
def Redo_recon(name=None):
    try:
        delete_reconcilation = delete_recon(
            {"recon_ids": [name], "redo": True})
        if not delete_reconcilation["success"]:
            return delete_reconcilation
        create_recon = create_reconciliation(redo=True, recon_id=name)
        if not create_recon["success"]:
            return create_recon
        return {"success": True, "message": "Recon Recreated"}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("Redo_recon",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}

@frappe.whitelist(allow_guest=True)
def outward_supply(month=None, year=None):
    try:
        start_date = year+'-'+month+"-01"
        end_date = str(date_util.get_last_day(start_date))
        outward = execute({"from_date": start_date, "to_date": end_date, "export": True})
        return outward
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("outward_supply",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}

@frappe.whitelist(allow_guest=True)
def last_recon_update(month=None, year=None):
    try:
        # if not month or not year:
        #     return {"success": False, "message": "month or year is mandatory"}
        recondate = ""
        # outward = frappe.db.sql("""SELECT `tabRecon Details`.creation as Creation from `tabRecon Details` where MONTH(`tabRecon Details`.start_date) = {} and YEAR(`tabRecon Details`.start_date) = {} order by creation desc LIMIT  1""".format(month, year), as_dict=1)
        outward = frappe.db.sql("""SELECT `tabRecon Details`.creation as Creation from `tabRecon Details` order by creation desc LIMIT  1""", as_dict=1)
        if len(outward) > 0:
            recondate = outward[0]["Creation"]
        return {"success": True, "date": recondate}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("last_recon_update",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}

@frappe.whitelist(allow_guest=True)
def dashborad_recon(recon_id=None):
    try:
        invoice_count = frappe.db.sql("""SELECT ezyinvoicing_count as 'ezyinvoicing_count', missing_in_ezyinvoicing as 'missing_in_ezyinvoicing', opera_folios_count as 'opera_folio_count', missing_in_opera as 'missing_in_opera', (ezyinvoicing_count-opera_folios_count) as 'ezy_opera_variance' from `tabInvoice Count` where recon_id='{}'""".format(recon_id), as_dict=1)
        ezy_opera_comparison = frappe.db.sql("""SELECT round(sum(ezyinvoicingbaseamount),2) as 'ezy_total_base_amount', round(sum(ezyinvoicinginvoiceamount),2) as 'ezy_total_invoice_amount', round(sum(operabaseamount),2) as 'opera_total_base_amount', round(sum(operainvoiceamount),2) as 'opera_total_invoice_amount' from `tabRecon Opera Comparison` where recon_id = '{}'""".format(recon_id), as_dict=1)
        con_b2c_b2b = frappe.db.sql("""SELECT count(invoicenumber) as 'count_of_b2c_to_b2b' from `tabConverted B2C to B2B` where recon_id = '{}'""".format(recon_id), as_dict=1)
        con_b2b_b2c = frappe.db.sql("""SELECT count(invoicenumber) as 'count_of_b2b_to_b2c' from `tabConverted B2B to B2C` where recon_id = '{}'""".format(recon_id), as_dict=1)
        opera_tax_credit = frappe.db.sql("""SELECT count(opera) as 'opera_count_tax_to_credit' from `tabInvoice Type Missmatch` where `tabInvoice Type Missmatch`.opera = "Tax Invoice" and recon_id = '{}'""".format(recon_id), as_dict=1)
        opera_credit_tax = frappe.db.sql("""SELECT count(opera) as 'opera_count_tax_to_credit' from `tabInvoice Type Missmatch` where `tabInvoice Type Missmatch`.opera = "Credit Invoice" and recon_id = '{}'""".format(recon_id), as_dict=1)
        # return opera_credit_tax
        total_data = {"invoice_count": {k: (0 if v is None else v) for k, v in invoice_count[0].items()},
                      "ezy_opera_comparison": {k: (0 if v is None else v) for k, v in ezy_opera_comparison[0].items()},
                      "con_b2c_b2b": {k: (0 if v is None else v) for k, v in con_b2c_b2b[0].items()},
                      "con_b2b_b2c": {k: (0 if v is None else v) for k, v in con_b2b_b2c[0].items()},
                      "opera_tax_credit": {k: (0 if v is None else v) for k, v in opera_tax_credit[0].items()},
                      "opera_credit_tax": {k: (0 if v is None else v) for k, v in opera_credit_tax[0].items()}}
        return {"success": True, "data": total_data}
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        frappe.log_error("last_recon_update",
                         "line No:{}\n{}".format(exc_tb.tb_lineno, str(e)))
        return {"success": False, "message": str(e)}


